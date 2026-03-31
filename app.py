"""
Content Quality Firewall — Web Interface
FastAPI backend serving the review engine + static HTML frontend.

Run:  python app.py
Then: open http://localhost:8500
"""

from __future__ import annotations

import json
import logging
import os
import re
import gzip
import asyncio
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlparse, urljoin
from collections import Counter

import httpx
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, FileResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles

from models.brand_context import BrandContext, ContentBrief
from firewall import ContentFirewall
from parsers.file_parser import parse_file
from outputs.root_cause import RootCauseDashboard
from engine.llm_judge import llm_review, set_model
from config.settings import (
    OPENAI_API_KEY,
    ANTHROPIC_API_KEY,
    DEFAULT_MODEL,
    MODEL_REGISTRY,
)

app = FastAPI(title="Vizup Soul — Content Quality Firewall")
app.mount("/static", StaticFiles(directory="static"), name="static")

log = logging.getLogger("vizup")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(message)s")

EXAMPLES_DIR = Path("examples")
HISTORY_PATH = "review_history.jsonl"

_KEY_CHECK = {
    "OPENAI_API_KEY": bool(OPENAI_API_KEY),
    "ANTHROPIC_API_KEY": bool(ANTHROPIC_API_KEY),
}

ALLOWED_EXTENSIONS = {".md", ".txt", ".html", ".htm", ".docx", ".pdf"}


@app.get("/", response_class=HTMLResponse)
async def index():
    return FileResponse("static/index.html")


@app.get("/api/models")
async def get_models():
    models = []
    for m in MODEL_REGISTRY:
        available = _KEY_CHECK.get(m["api_key_env"], False)
        models.append({
            "id": m["id"],
            "name": m["name"],
            "provider": m["provider"],
            "available": available,
        })
    default = DEFAULT_MODEL
    if not any(m["id"] == default and m["available"] for m in models):
        for m in models:
            if m["available"]:
                default = m["id"]
                break
    return {"models": models, "default": default}


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Parse an uploaded file and return extracted text."""
    filename = file.filename or "unknown.txt"
    ext = Path(filename).suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: {ext}. Supported: {', '.join(sorted(ALLOWED_EXTENSIONS))}",
        )

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="File is empty.")

    try:
        content = parse_file(file_bytes, filename)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to parse file: {e}")

    word_count = len(content.split())
    return {
        "content": content,
        "filename": filename,
        "file_type": ext,
        "word_count": word_count,
        "char_count": len(content),
    }


@app.post("/api/fetch-url")
async def fetch_url_content(url: str = Form(...)):
    """Fetch a URL and return extracted text as reviewable content."""
    try:
        page_text, word_count = await _fetch_and_extract(url)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to fetch URL: {e}")

    if word_count < 20:
        raise HTTPException(status_code=422, detail="Could not extract meaningful content from the URL.")

    return {
        "content": page_text,
        "word_count": word_count,
        "source_url": url,
    }


@app.get("/api/sample-brand")
async def get_sample_brand():
    path = EXAMPLES_DIR / "sample_brand_context.json"
    if path.exists():
        return json.loads(path.read_text())
    return {}


@app.get("/api/sample-brief")
async def get_sample_brief():
    path = EXAMPLES_DIR / "sample_brief.json"
    if path.exists():
        return json.loads(path.read_text())
    return {}


@app.get("/api/sample-content")
async def get_sample_content():
    path = EXAMPLES_DIR / "sample_content.md"
    if path.exists():
        return {"content": path.read_text()}
    return {"content": ""}


@app.post("/api/review")
async def run_review(
    content: str = Form(...),
    brand_context_json: str = Form(...),
    brief_json: str = Form(""),
    mode: str = Form("full"),
    model: str = Form(""),
):
    try:
        brand_data = json.loads(brand_context_json)
        brand = BrandContext(**brand_data)
    except (json.JSONDecodeError, Exception):
        brand = BrandContext(raw_description=brand_context_json.strip())

    brief = None
    if brief_json and brief_json.strip():
        try:
            brief_data = json.loads(brief_json)
            brief = ContentBrief(**brief_data)
        except (json.JSONDecodeError, Exception):
            brief = ContentBrief(notes=brief_json.strip())

    firewall = ContentFirewall(
        brand_context=brand,
        brief=brief,
        history_path=HISTORY_PATH,
    )

    model_to_use = model if model else DEFAULT_MODEL

    if mode == "quick":
        verdict = await firewall.quick_review(content, model=model_to_use)
    else:
        verdict = await firewall.review(content, model=model_to_use)

    redline_json = firewall.get_redline_json()

    repairs = None
    if mode == "audit" and verdict.decision.value != "pass":
        repairs = await firewall.get_repairs()

    return {
        "decision": verdict.decision.value,
        "overall_score": verdict.overall_score,
        "total_issues": verdict.total_issues,
        "blocker_count": verdict.blocker_count,
        "major_count": verdict.major_count,
        "minor_count": verdict.minor_count,
        "style_count": verdict.style_count,
        "issues": redline_json.get("issues", []),
        "score_breakdown": redline_json.get("score_breakdown", {}),
        "reviewer_summaries": redline_json.get("reviewer_summaries", {}),
        "repairs": repairs,
        "model_used": model_to_use,
    }


@app.get("/api/dashboard")
async def get_dashboard():
    dashboard = RootCauseDashboard(history_path=HISTORY_PATH)
    if not dashboard._history:
        return {"text": "No review history yet. Run some reviews first.", "data": None}

    text = dashboard.generate_dashboard()
    trending = dashboard.get_trending_issues()

    category_counts: dict[str, int] = {}
    total = len(dashboard._history)
    for entry in dashboard._history[-50:]:
        for cat in set(entry.get("categories", [])):
            category_counts[cat] = category_counts.get(cat, 0) + 1

    return {
        "text": text,
        "data": {
            "total_reviews": total,
            "category_rates": {
                k: round(v / min(total, 50) * 100, 1)
                for k, v in sorted(category_counts.items(), key=lambda x: -x[1])
            },
            "trending": trending,
        },
    }


@app.post("/api/competitor-analysis")
async def competitor_analysis(
    url: str = Form(...),
    model: str = Form(""),
):
    """Fetch a competitor URL, extract content, and analyze it with LLM."""
    model_to_use = model if model else DEFAULT_MODEL
    set_model(model_to_use)

    try:
        page_text, word_count = await _fetch_and_extract(url)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to fetch URL: {e}")

    if word_count < 50:
        raise HTTPException(status_code=422, detail="Could not extract meaningful content from URL.")

    max_chars = 12000
    analysis_text = page_text[:max_chars]

    system_prompt = """You are an expert content strategist and SEO analyst. You analyze competitor content that is currently ranking on Google to understand search intent, writing quality, and content patterns.

Return a JSON object with these exact keys:
{
  "search_intent": "What search intent does this content serve? What questions does it answer? What is Google rewarding here? What does the reader expect when they click this? Be specific and actionable.",
  "writing_style": "Analyze the writing style: tone (formal/casual/technical), sentence structure (short/long/mixed), use of data/examples/stories, persuasion techniques, how they handle transitions, paragraph length patterns, use of active vs passive voice. Note specific patterns.",
  "structure_analysis": "How is the content structured? Heading hierarchy, section flow, use of lists/tables/visuals, intro pattern, conclusion pattern, CTA placement. What structural choices make this rank?",
  "ai_analysis": "Is this likely AI-written or human-written? Look for: repetitive sentence openers, generic frameworks, lack of specific examples/data, formulaic transitions, semantic repetition, overly balanced 'on one hand / on the other' patterns, absence of genuine opinion or expertise signals. Give specific evidence.",
  "ai_probability": 0.0 to 1.0,
  "reading_level": "Grade level or complexity descriptor",
  "recommendations": "Based on this analysis, what should the user do differently in their own content to compete? Be specific: what to copy, what to avoid, what gaps to exploit, what angle to take."
}"""

    user_prompt = f"""Analyze this competitor content that is currently ranking on Google.

URL: {url}
Word count: {word_count}

--- CONTENT ---
{analysis_text}
--- END ---

Provide your analysis as JSON."""

    try:
        result = await llm_review(
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            model=model_to_use,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LLM analysis failed: {e}")

    if result.get("parse_error"):
        raise HTTPException(status_code=500, detail="Failed to parse LLM response.")

    result["word_count"] = word_count
    result["url"] = url
    result["model_used"] = model_to_use
    return result


@app.post("/api/generate-content-brief")
async def generate_content_brief(
    competitor_url: str = Form(""),
    competitor_content: str = Form(""),
    competitor_analysis_json: str = Form(""),
    brand_context_text: str = Form(""),
    model: str = Form(""),
):
    """Generate a structured content creation brief from competitor analysis."""
    model_to_use = model if model else DEFAULT_MODEL
    set_model(model_to_use)

    system_prompt = """You are an elite content strategist who creates detailed content creation briefs. You analyze competitor content that is currently ranking and reverse-engineer the exact instructions needed to create superior content.

You MUST return a JSON object with ALL of these keys. Every field must be filled with specific, actionable recommendations — never leave a field empty or say "N/A":

{
  "page_kind": "blog_post" or "landing_page",
  "page_kind_reasoning": "Why this page kind is recommended",

  "content_goal": exactly one of: "educate_inform", "thought_leadership", "drive_organic_traffic", "showcase_results", "compare_evaluate", "generate_leads",
  "content_goal_reasoning": "Why this goal fits best",

  "primary_keyword": "The exact primary keyword/topic to target",
  "primary_keyword_reasoning": "Why this keyword and how the competitor is using it",

  "content_nature": exactly one of: "how_to", "what_is", "listicle", "comparison", "guide", "case_study", "review", "opinion", "alternative",
  "content_nature_reasoning": "Why this content structure fits",

  "headline_recommendations": ["5 specific headline options, ranked by predicted CTR"],

  "secondary_keywords": ["6-8 secondary keywords extracted from competitor + gaps identified"],

  "tonality": exactly one of: "professional", "conversational", "technical", "founder_voice", "authoritative", "friendly_approachable",
  "tonality_reasoning": "Why this tone works for this topic/audience",

  "theme_category": "The theme or category this content belongs to",

  "depth_word_count": exactly one of: "short_500_800", "medium_1000_1500", "long_1500_3000", "deep_dive_3000_plus",
  "depth_reasoning": "Why this word count range based on competitor analysis",

  "reading_format": exactly one of: "narrative", "step_by_step", "qa_format", "hybrid",
  "reading_format_reasoning": "Why this reading format",

  "components": {
    "include_faq": true/false,
    "faq_reasoning": "Why or why not, and suggested FAQ topics if yes",
    "include_table": true/false,
    "table_reasoning": "What tables/comparison matrices to include",
    "include_embeds": true/false,
    "embed_reasoning": "What embeds would strengthen this content",
    "include_proof_block": true/false,
    "proof_reasoning": "What proof elements to include",
    "include_callouts": true/false,
    "callout_reasoning": "What callout types and content",
    "include_ctas": true/false,
    "cta_reasoning": "CTA placement strategy and copy direction",
    "include_images": true/false,
    "image_reasoning": "What images/visuals to include and where"
  },

  "custom_persona": "A specific writing persona description tailored for this content — who the writer should be, their expertise, their voice",

  "custom_instructions": "Detailed paragraph of specific instructions for the content writer. Include: sections to cover, angles to emphasize, competitor gaps to exploit, specific data points to include, things the competitor does well to match, things they do poorly to beat. Be very specific and actionable.",

  "suggested_outline": ["Ordered list of section headings that form the article outline"],

  "competitor_gaps": ["Specific things the competitor missed or did poorly that we should capitalize on"],

  "differentiation_strategy": "How to make this content clearly better and different from the competitor"
}"""

    user_prompt_parts = []

    if competitor_url:
        user_prompt_parts.append(f"Competitor URL: {competitor_url}")

    if competitor_analysis_json:
        try:
            analysis = json.loads(competitor_analysis_json)
            user_prompt_parts.append(f"Competitor Analysis Summary:\n- Search Intent: {analysis.get('search_intent', '')}\n- Writing Style: {analysis.get('writing_style', '')}\n- Structure: {analysis.get('structure_analysis', '')}\n- AI Analysis: {analysis.get('ai_analysis', '')}\n- AI Probability: {analysis.get('ai_probability', '')}\n- Recommendations: {analysis.get('recommendations', '')}")
        except json.JSONDecodeError:
            pass

    if brand_context_text:
        user_prompt_parts.append(f"Brand Context:\n{brand_context_text}")

    if competitor_content:
        max_chars = 10000
        user_prompt_parts.append(f"--- COMPETITOR CONTENT (first {max_chars} chars) ---\n{competitor_content[:max_chars]}\n--- END ---")

    user_prompt = "\n\n".join(user_prompt_parts)
    user_prompt += "\n\nBased on the above, generate a complete content creation brief as JSON. Every field must be filled with specific, actionable recommendations."

    try:
        result = await llm_review(
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            model=model_to_use,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LLM brief generation failed: {e}")

    if result.get("parse_error"):
        raise HTTPException(status_code=500, detail="Failed to parse LLM response for content brief.")

    result["model_used"] = model_to_use
    return result


async def _fetch_and_extract(url: str) -> tuple[str, int]:
    """Fetch a URL and return (extracted_text, word_count)."""
    async with httpx.AsyncClient(
        follow_redirects=True,
        timeout=30.0,
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/131.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    ) as client:
        resp = await client.get(url)
        resp.raise_for_status()

    from bs4 import BeautifulSoup, NavigableString, Tag

    soup = BeautifulSoup(resp.text, "html.parser")
    for tag in soup(["script", "style", "noscript", "iframe", "svg"]):
        tag.decompose()

    # Try to narrow to the main content area
    MIN_CONTENT_LEN = 500

    def _has_enough_content(tag):
        return tag is not None and len(tag.get_text(strip=True)) >= MIN_CONTENT_LEN

    def _find_content_div_by_class(root):
        keywords = (
            "post-content", "article-body", "entry-content",
            "blog-content", "page-content", "content-body",
            "markdown-body", "prose", "blog-post", "blog_post",
            "section_blog", "post-body", "article-content",
            "rich-text", "richtext", "wysiwyg",
        )
        for kw in keywords:
            found = root.find(
                ["div", "section"],
                class_=lambda c: c and kw in (c if isinstance(c, str) else " ".join(c)).lower(),
            )
            if _has_enough_content(found):
                return found
        return None

    def _find_largest_content_block(root):
        """Find the deepest div/section that contains the most paragraph-rich content."""
        best, best_score = None, 0
        for el in root.find_all(["div", "section"]):
            p_count = len(el.find_all("p", recursive=False)) + len(el.find_all("p"))
            text_len = len(el.get_text(strip=True))
            score = p_count * 100 + text_len
            if score > best_score and text_len >= MIN_CONTENT_LEN:
                best_score = score
                best = el
        return best

    import copy

    def _cleanup(root):
        """Remove nav/footer/sidebar noise from a content root (in-place)."""
        for tag in root.find_all(["nav", "footer", "aside"]):
            tag.decompose()
        for tag in root.find_all(
            attrs={"class": lambda c: c and any(
                k in (c if isinstance(c, str) else " ".join(c)).lower()
                for k in ("sidebar", "related-posts", "share-buttons",
                           "social-share", "newsletter", "cookie", "popup",
                           "comments", "comment-form")
            )}
        ):
            tag.decompose()
        for tag in root.find_all("header"):
            if tag.find("nav") or tag.find(class_=lambda c: c and "logo" in str(c).lower()):
                tag.decompose()
        return root

    def _try_root(candidate):
        """Copy, clean, and check if a candidate root has enough content."""
        if candidate is None:
            return None
        trial = copy.copy(candidate)
        _cleanup(trial)
        if len(trial.get_text(strip=True)) >= MIN_CONTENT_LEN:
            _cleanup(candidate)
            return candidate
        return None

    content_root = None

    # 1. Semantic tags (article, main)
    for c in [soup.find("article"), soup.find("main"), soup.find("div", role="main")]:
        content_root = _try_root(c)
        if content_root:
            break

    # 2. Class-based content divs
    if content_root is None:
        found = _find_content_div_by_class(soup)
        content_root = _try_root(found)

    # 3. Largest content-rich block
    if content_root is None:
        found = _find_largest_content_block(soup.body or soup)
        content_root = _try_root(found)

    # 4. Fallback to body
    if content_root is None:
        content_root = soup.body or soup
        _cleanup(content_root)

    seen_texts = set()
    blocks = []

    def _add(text: str, prefix: str = ""):
        text = text.strip()
        if not text:
            return
        line = f"{prefix}{text}" if prefix else text
        # Deduplicate identical blocks (nested elements can repeat)
        sig = text[:120]
        if sig in seen_texts:
            return
        seen_texts.add(sig)
        blocks.append(line)

    def _table_to_markdown(table_tag) -> str:
        """Convert an HTML table to a markdown table."""
        rows = []
        for tr in table_tag.find_all("tr"):
            cells = []
            for td in tr.find_all(["th", "td"]):
                cells.append(td.get_text(strip=True).replace("|", "\\|"))
            if cells:
                rows.append(cells)
        if not rows:
            return ""
        # Normalize column count
        max_cols = max(len(r) for r in rows)
        for r in rows:
            while len(r) < max_cols:
                r.append("")
        lines = ["| " + " | ".join(rows[0]) + " |"]
        lines.append("| " + " | ".join(["---"] * max_cols) + " |")
        for r in rows[1:]:
            lines.append("| " + " | ".join(r) + " |")
        return "\n".join(lines)

    # Remove common boilerplate elements before walking
    for tag in content_root.find_all(
        attrs={"class": lambda c: c and any(
            k in (c if isinstance(c, str) else " ".join(c)).lower()
            for k in ("related-post", "also-like", "you-may-like", "recommended",
                       "share-bar", "social-bar", "author-bio", "author-box",
                       "breadcrumb", "pagination", "tag-list", "category-list",
                       "ad-slot", "advertisement", "promo-banner", "cta-block",
                       "sidebar-widget", "widget-area", "also_like", "related_post",
                       "wp-block-group", "infobox", "callout-box")
        )}
    ):
        tag.decompose()

    # Remove "You may also like" sections (heading + its sibling content)
    for tag in content_root.find_all(string=lambda s: s and "you may also like" in s.lower()):
        parent = tag.find_parent(["div", "section", "aside"])
        if parent and len(parent.get_text(strip=True)) < 500:
            parent.decompose()

    # Walk the DOM tree in document order
    for el in content_root.descendants:
        if not isinstance(el, Tag):
            continue

        # Skip elements nested inside tags we'll handle at their level
        if el.find_parent(["table", "pre"]) and el.name not in ("table", "pre"):
            continue

        if el.name in ("h1", "h2", "h3", "h4", "h5", "h6"):
            level = int(el.name[1])
            _add(el.get_text(strip=True), "#" * level + " ")

        elif el.name == "p":
            _add(el.get_text(strip=True))

        elif el.name == "blockquote":
            _add(el.get_text(strip=True), "> ")

        elif el.name == "li":
            _add(el.get_text(strip=True), "- ")

        elif el.name == "dt":
            _add(el.get_text(strip=True), "**")
            if blocks and blocks[-1].startswith("**"):
                blocks[-1] += "**"

        elif el.name == "dd":
            _add(el.get_text(strip=True), "  ")

        elif el.name == "table":
            md_table = _table_to_markdown(el)
            if md_table:
                sig = md_table[:120]
                if sig not in seen_texts:
                    seen_texts.add(sig)
                    blocks.append(md_table)

        elif el.name in ("pre", "code"):
            if el.name == "code" and el.find_parent("pre"):
                continue
            code_text = el.get_text(strip=False).strip()
            if code_text and len(code_text) > 3:
                sig = code_text[:120]
                if sig not in seen_texts:
                    seen_texts.add(sig)
                    blocks.append(f"```\n{code_text}\n```")

        elif el.name in ("figcaption",):
            _add(el.get_text(strip=True), "[Caption: ", )
            if blocks and blocks[-1].startswith("[Caption: "):
                blocks[-1] += "]"

        elif el.name == "img":
            alt = el.get("alt", "").strip()
            # Only keep meaningful alt text (skip filenames and slugs)
            if alt and len(alt) > 20 and not alt.replace("-", " ").replace("_", " ").islower():
                _add(f"[Image: {alt}]")

        elif el.name == "div":
            # Only capture divs that contain direct text (not just wrapper divs)
            direct_text = "".join(
                child.strip() for child in el.children
                if isinstance(child, NavigableString) and child.strip()
            )
            if len(direct_text) > 30:
                _add(direct_text)

    # Post-process: remove boilerplate lines
    NOISE_PATTERNS = [
        "alternatively, book", "book with free", "check rates and book",
        "book via", "official website", "photos ©", "photos (c)",
        "see what others think", "or check tripadvisor",
        "subscribe to", "newsletter", "follow us",
        "facebook twitter", "pinterest instagram",
        "we will never spam", "we will never share",
        "cookie", "© travel", "site map",
        "this post contains affiliate", "this post includes affiliate",
        "we may earn a commission", "at no extra cost",
        "media / pr", "contact us:", "about us",
    ]
    cleaned_blocks = []
    for block in blocks:
        bl = block.lower().strip()
        if any(p in bl for p in NOISE_PATTERNS):
            continue
        # Skip very short blocks that are just link labels
        if len(bl) < 15 and any(k in bl for k in [
            "booking.com", "expedia", "tripadvisor", "agoda",
            "hotels.com", "tablet hotels", "mr & mrs smith",
            "read more", "learn more", "view all",
        ]):
            continue
        cleaned_blocks.append(block)

    page_text = "\n\n".join(cleaned_blocks)

    # Fallback: if extraction got very little, use full text
    if len(page_text.split()) < 50:
        page_text = content_root.get_text(separator="\n", strip=True)

    return page_text, len(page_text.split())


def _extract_page_context(html: str, url: str) -> dict:
    """Extract structural and contextual metadata from a page's HTML."""
    from bs4 import BeautifulSoup, Tag
    from urllib.parse import urlparse

    soup = BeautifulSoup(html, "html.parser")
    domain = urlparse(url).netloc

    from urllib.parse import unquote

    # Meta tags
    title_tag = soup.find("title")
    meta_desc = soup.find("meta", attrs={"name": "description"})
    og_title = soup.find("meta", attrs={"property": "og:title"})
    og_desc = soup.find("meta", attrs={"property": "og:description"})
    og_image = soup.find("meta", attrs={"property": "og:image"})
    canonical = soup.find("link", attrs={"rel": "canonical"})
    robots_meta = soup.find("meta", attrs={"name": "robots"})

    # SERP-focused analysis
    parsed_url = urlparse(url)
    url_slug = unquote(parsed_url.path.rstrip("/").split("/")[-1]) if parsed_url.path else ""
    url_path = unquote(parsed_url.path)
    url_depth = len([p for p in parsed_url.path.strip("/").split("/") if p])
    title_text = title_tag.get_text(strip=True) if title_tag else ""
    title_len = len(title_text)
    desc_text = meta_desc["content"].strip() if meta_desc and meta_desc.get("content") else ""
    desc_len = len(desc_text)

    # Heading structure
    heading_counts = {}
    heading_outline = []
    for level in range(1, 7):
        tag_name = f"h{level}"
        tags = soup.find_all(tag_name)
        heading_counts[tag_name] = len(tags)
        for t in tags[:20]:
            text = t.get_text(strip=True)[:80]
            if text:
                heading_outline.append(f"{'  ' * (level - 1)}{tag_name}: {text}")

    # Links analysis — deep extraction
    all_links = soup.find_all("a", href=True)
    internal_links = 0
    external_links = 0
    external_domains = set()
    nofollow_count = 0
    sponsored_count = 0
    internal_anchors = []
    external_link_details = []

    content_area = (
        soup.find("article") or soup.find("main")
        or soup.find("div", role="main") or soup.body or soup
    )

    for a in all_links:
        href = a["href"]
        if href.startswith("#") or href.startswith("javascript:") or href.startswith("mailto:"):
            continue
        parsed = urlparse(href)
        anchor = a.get_text(strip=True)[:80]
        rel = " ".join(a.get("rel", []))
        is_nofollow = "nofollow" in rel
        is_sponsored = "sponsored" in rel or "ugc" in rel
        in_content = bool(a.find_parent(content_area.name) if content_area else False)

        if is_nofollow:
            nofollow_count += 1
        if is_sponsored:
            sponsored_count += 1

        if not parsed.netloc or parsed.netloc == domain:
            internal_links += 1
            if anchor and len(anchor) > 2 and in_content:
                internal_anchors.append(anchor)
        else:
            external_links += 1
            external_domains.add(parsed.netloc)
            if anchor and len(anchor) > 2:
                external_link_details.append({
                    "domain": parsed.netloc,
                    "anchor": anchor,
                    "nofollow": is_nofollow,
                    "sponsored": is_sponsored,
                })

    # Categorize external domains
    ext_domain_counts = {}
    for d in [el["domain"] for el in external_link_details]:
        ext_domain_counts[d] = ext_domain_counts.get(d, 0) + 1
    top_ext_domains_with_count = sorted(ext_domain_counts.items(), key=lambda x: -x[1])[:15]

    # Images
    all_images = soup.find_all("img")
    images_with_alt = sum(1 for img in all_images if img.get("alt", "").strip())

    # Content structure elements
    body = soup.body or soup
    total_lists = len(body.find_all(["ul", "ol"]))
    total_tables = len(body.find_all("table"))
    total_blockquotes = len(body.find_all("blockquote"))
    total_code_blocks = len(body.find_all(["pre", "code"]))
    total_videos = len(body.find_all(["video", "iframe"]))

    # E-E-A-T signals
    has_author = bool(
        soup.find(attrs={"class": lambda c: c and "author" in str(c).lower()})
        or soup.find("meta", attrs={"name": "author"})
        or soup.find(attrs={"rel": "author"})
    )
    has_date = bool(
        soup.find("time")
        or soup.find(attrs={"class": lambda c: c and any(
            k in str(c).lower() for k in ("date", "publish", "posted", "updated")
        )})
        or soup.find("meta", attrs={"property": "article:published_time"})
    )
    has_schema = bool(
        soup.find("script", attrs={"type": "application/ld+json"})
    )

    # Schema details
    schema_types = []
    def _collect_schema_type(val):
        if isinstance(val, str):
            schema_types.append(val)
        elif isinstance(val, list):
            for v in val:
                if isinstance(v, str):
                    schema_types.append(v)

    for script in soup.find_all("script", attrs={"type": "application/ld+json"}):
        try:
            sd = json.loads(script.string or "")
            if isinstance(sd, dict) and "@type" in sd:
                _collect_schema_type(sd["@type"])
            elif isinstance(sd, list):
                for item in sd:
                    if isinstance(item, dict) and "@type" in item:
                        _collect_schema_type(item["@type"])
        except (json.JSONDecodeError, TypeError):
            pass

    # Navigation/TOC detection
    has_toc = bool(
        soup.find(attrs={"class": lambda c: c and any(
            k in str(c).lower() for k in ("toc", "table-of-contents", "tableofcontents")
        )})
        or soup.find(attrs={"id": lambda i: i and "toc" in str(i).lower()})
    )

    # FAQ section
    has_faq = bool(
        soup.find(string=lambda s: s and "frequently asked" in s.lower())
        or soup.find(attrs={"class": lambda c: c and "faq" in str(c).lower()})
        or any(st == "FAQPage" for st in schema_types)
    )

    # Social proof / engagement signals
    has_social_share = bool(
        soup.find(attrs={"class": lambda c: c and any(
            k in str(c).lower() for k in ("share", "social")
        )})
    )
    has_comments = bool(
        soup.find(attrs={"class": lambda c: c and "comment" in str(c).lower()})
        or soup.find(attrs={"id": lambda i: i and "comment" in str(i).lower()})
    )

    return {
        # SERP / CTR signals
        "title_tag": title_text,
        "title_length": title_len,
        "title_pixel_ok": title_len <= 60,
        "meta_description": desc_text,
        "meta_description_length": desc_len,
        "meta_description_pixel_ok": 120 <= desc_len <= 160,
        "og_title": (og_title["content"] if og_title and og_title.get("content") else ""),
        "og_description": (og_desc["content"] if og_desc and og_desc.get("content") else ""),
        "og_image": bool(og_image),
        "canonical_url": (canonical["href"] if canonical and canonical.get("href") else ""),
        "robots_meta": (robots_meta["content"] if robots_meta and robots_meta.get("content") else ""),
        "url_slug": url_slug,
        "url_path": url_path,
        "url_depth": url_depth,
        # Structure
        "heading_counts": heading_counts,
        "heading_outline": heading_outline[:30],
        "total_images": len(all_images),
        "images_with_alt": images_with_alt,
        "internal_links": internal_links,
        "internal_anchor_samples": internal_anchors[:20],
        "external_links": external_links,
        "nofollow_links": nofollow_count,
        "sponsored_links": sponsored_count,
        "top_external_domains": [f"{d} ({c})" for d, c in top_ext_domains_with_count],
        "external_link_samples": external_link_details[:15],
        "unique_external_domains": len(external_domains),
        "total_lists": total_lists,
        "total_tables": total_tables,
        "total_blockquotes": total_blockquotes,
        "total_code_blocks": total_code_blocks,
        "total_videos": total_videos,
        # E-E-A-T
        "has_author": has_author,
        "has_published_date": has_date,
        "has_schema_markup": has_schema,
        "schema_types": schema_types[:10],
        "has_toc": has_toc,
        "has_faq": has_faq,
        "has_social_share": has_social_share,
        "has_comments": has_comments,
    }


async def _fetch_page_html(url: str) -> str:
    """Fetch a URL and return raw HTML."""
    async with httpx.AsyncClient(
        follow_redirects=True,
        timeout=30.0,
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/131.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    ) as client:
        resp = await client.get(url)
        resp.raise_for_status()
    return resp.text


def _format_context_for_prompt(ctx: dict) -> str:
    """Format page context metadata into a concise text block for LLM analysis."""
    if not ctx:
        return "No context data available."
    lines = []

    # SERP Listing section
    lines.append("-- SERP LISTING --")
    if ctx.get("title_tag"):
        ok = "OK" if ctx.get("title_pixel_ok") else "TOO LONG"
        lines.append(f"Title Tag ({ctx.get('title_length', '?')} chars, {ok}): {ctx['title_tag']}")
    if ctx.get("meta_description"):
        ok = "OK" if ctx.get("meta_description_pixel_ok") else ("TOO SHORT" if ctx.get("meta_description_length", 0) < 120 else "TOO LONG")
        lines.append(f"Meta Description ({ctx.get('meta_description_length', '?')} chars, {ok}): {ctx['meta_description'][:200]}")
    if ctx.get("url_slug"):
        lines.append(f"URL Slug: {ctx['url_slug']}")
    if ctx.get("url_path"):
        lines.append(f"URL Path (depth {ctx.get('url_depth', '?')}): {ctx['url_path']}")
    if ctx.get("canonical_url"):
        lines.append(f"Canonical URL: {ctx['canonical_url']}")
    og_parts = []
    if ctx.get("og_title"):
        og_parts.append(f"title='{ctx['og_title'][:60]}'")
    if ctx.get("og_description"):
        og_parts.append(f"desc='{ctx['og_description'][:60]}...'")
    og_parts.append("image=YES" if ctx.get("og_image") else "image=NO")
    lines.append(f"Open Graph: {', '.join(og_parts)}")
    if ctx.get("robots_meta"):
        lines.append(f"Robots Meta: {ctx['robots_meta']}")

    # Page structure section
    lines.append("-- PAGE STRUCTURE --")
    hc = ctx.get("heading_counts", {})
    if hc:
        h_str = ", ".join(f"{k}: {v}" for k, v in hc.items() if v > 0)
        lines.append(f"Heading Structure: {h_str}")
    lines.append(f"Images: {ctx.get('total_images', 0)} total ({ctx.get('images_with_alt', 0)} with alt text)")

    # Dedicated linking section
    lines.append("-- LINKING PROFILE --")
    lines.append(f"Internal Links: {ctx.get('internal_links', 0)}")
    if ctx.get("internal_anchor_samples"):
        samples = ctx["internal_anchor_samples"][:10]
        lines.append(f"Internal Anchor Samples: {' | '.join(samples)}")
    lines.append(f"External Links: {ctx.get('external_links', 0)} (to {ctx.get('unique_external_domains', 0)} unique domains)")
    if ctx.get("nofollow_links"):
        lines.append(f"Nofollow Links: {ctx['nofollow_links']}")
    if ctx.get("sponsored_links"):
        lines.append(f"Sponsored/UGC Links: {ctx['sponsored_links']}")
    if ctx.get("top_external_domains"):
        lines.append(f"Top External Domains: {', '.join(ctx['top_external_domains'][:10])}")
    if ctx.get("external_link_samples"):
        ext_samples = []
        for el in ctx["external_link_samples"][:8]:
            nf = " [nofollow]" if el.get("nofollow") else ""
            sp = " [sponsored]" if el.get("sponsored") else ""
            ext_samples.append(f"'{el['anchor']}' -> {el['domain']}{nf}{sp}")
        lines.append(f"External Link Details: {' | '.join(ext_samples)}")
    elements = []
    if ctx.get("total_lists"):
        elements.append(f"{ctx['total_lists']} lists")
    if ctx.get("total_tables"):
        elements.append(f"{ctx['total_tables']} tables")
    if ctx.get("total_blockquotes"):
        elements.append(f"{ctx['total_blockquotes']} blockquotes")
    if ctx.get("total_code_blocks"):
        elements.append(f"{ctx['total_code_blocks']} code blocks")
    if ctx.get("total_videos"):
        elements.append(f"{ctx['total_videos']} videos/embeds")
    if elements:
        lines.append(f"Rich Elements: {', '.join(elements)}")
    trust = []
    if ctx.get("has_author"):
        trust.append("Author attribution")
    if ctx.get("has_published_date"):
        trust.append("Publish date")
    if ctx.get("has_schema_markup"):
        types = ctx.get("schema_types", [])
        trust.append(f"Schema markup ({', '.join(types[:3])})" if types else "Schema markup")
    if ctx.get("has_toc"):
        trust.append("Table of contents")
    if ctx.get("has_faq"):
        trust.append("FAQ section")
    if ctx.get("has_comments"):
        trust.append("Comments section")
    if ctx.get("has_social_share"):
        trust.append("Social share buttons")
    if trust:
        lines.append(f"Trust & UX Signals: {', '.join(trust)}")
    return "\n".join(lines)


def _build_ranking_analysis_prompt(main_data, comp_data, page_blocks, has_main):
    """Build prompts that learn from what Google ranks, then evaluate the main page."""
    all_pages = ([main_data] if main_data else []) + list(comp_data)
    n = len(all_pages)
    labels = [p["label"] for p in all_pages]

    article_entries = []
    for p in all_pages:
        pos = p.get("position", "")
        article_entries.append(
            f'    {{ "label": "{p["label"]}", "url": "{p["url"]}", '
            f'"word_count": {p["word_count"]}, '
            f'"ranking_position": "{pos}", '
            f'"overall_grade": "A+ to F" }}'
        )
    articles_json = ",\n".join(article_entries)
    scores_csv = ", ".join(f"score_{i}" for i in range(n))

    analysis_fields = "\n      ".join(
        f'"{labels[i].lower().replace(" ", "_")}_analysis": "Analysis of {labels[i]}",'
        for i in range(n)
    )

    system_prompt = f"""You are a senior search strategist who reverse-engineers Google's ranking decisions. You do NOT do checklist SEO audits. Instead, you study what Google is actually ranking and learn its preferences.

YOUR METHOD:
1. STUDY the top-ranking pages (Page 1 competitors). These are what Google currently rewards.
2. EXTRACT PATTERNS from the ranking pages — what word count range does Google prefer? What heading density? What page type? What content structure? What trust signals? What intent match?
3. CALCULATE the "Google-preferred range" for each dimension based on what is ACTUALLY ranking.
4. EVALUATE the main page (Your Page) against these learned ranges — not against SEO best practices, but against what Google demonstrably rewards for this query.
5. IDENTIFY the real reasons for ranking gaps — not surface-level SEO features.

CRITICAL RULES:
- MORE is NOT better. 8000 words is NOT better than 2000 if the ranking pages average 2000.
- MORE headings is NOT better. 33 H2s can be worse than 8 if ranking pages have 8.
- MORE schema is NOT better. Triple schema on a page-45 result means nothing.
- Unicode tricks in titles are NOT ranking signals. Do not recommend them.
- OVER-OPTIMIZATION is a NEGATIVE signal. Detect and penalize it.
- Page type FIT matters enormously. A mega-guide loses to a clean package page if the query is transactional.
- Site-level trust drags individual pages down. Acknowledge when this is likely.
- Content EFFICIENCY (value per word) beats content LENGTH.

Return a JSON object:
{{
  "articles": [
{articles_json}
  ],

  "google_preference_model": {{
    "query_intent": "What does the searcher actually want? Be specific.",
    "preferred_page_type": "What page type does Google prefer for this query based on what ranks?",
    "preferred_word_count_range": "e.g. 1500-3000 based on what ranks",
    "preferred_heading_density": "e.g. 6-12 H2s, based on ranking pages",
    "preferred_content_style": "Direct/transactional vs comprehensive/editorial — what ranks?",
    "trust_signals_that_matter": "What trust elements do the TOP ranking pages share?",
    "what_ranking_pages_have_in_common": "3-5 patterns shared by Page 1 results",
    "what_ranking_pages_do_NOT_do": "Things the ranking pages avoid that weaker pages do"
  }},

  "overall_verdict": "3-5 sentences. What does Google prefer for this query based on the ranking evidence? If there is a main page: why is it ranked where it is? What is the #1 reason for the ranking gap? Be brutally honest.",

  "comparative_scores": {{
    "Query Intent Match": [{scores_csv}],
    "Content Efficiency": [{scores_csv}],
    "Page Type Fit": [{scores_csv}],
    "Trust & Authority Signals": [{scores_csv}],
    "User Decision Support": [{scores_csv}],
    "Content Focus vs Bloat": [{scores_csv}],
    "Conversion UX": [{scores_csv}],
    "Over-Optimization Risk": [{scores_csv}],
    "SERP Click Appeal": [{scores_csv}],
    "Overall Ranking Potential": [{scores_csv}]
  }},

  "section_comparisons": [
    {{
      "dimension": "Query Intent Alignment",
      {analysis_fields}
      "verdict": "Which page best matches what the searcher wants?",
      "google_signal": "What does the ranking data tell us about Google's intent interpretation?"
    }},
    {{ "dimension": "Content Efficiency & Focus", ... }},
    {{ "dimension": "Page Type & Structure Fit", ... }},
    {{ "dimension": "Over-Optimization Detection", ... }},
    {{ "dimension": "Trust, Authority & E-E-A-T", ... }},
    {{ "dimension": "User Decision Journey", ... }},
    {{ "dimension": "SERP Listing & Click Appeal", ... }},
    {{ "dimension": "Internal & External Linking Quality", ... }},
    {{ "dimension": "Content Uniqueness & Differentiation", ... }},
    {{ "dimension": "Conversion Path & CTA Clarity", ... }}
  ],

  "ranking_gap_diagnosis": {{
    "primary_reason": "The single most important reason for the ranking gap",
    "secondary_reasons": ["2-3 additional contributing factors"],
    "site_level_concerns": "Any site-level trust/quality issues that may be dragging this page down",
    "over_optimization_flags": ["Specific things the main page does that may be hurting it"],
    "intent_mismatch_details": "How the main page's approach differs from what Google rewards"
  }},

  "action_items": [
    "Specific, prioritized action. Focus on what will ACTUALLY improve rankings, not checklist items.",
    "... up to 10 actions, ordered by expected ranking impact"
  ]
}}

SCORING RULES:
- "Query Intent Match": How well does the page serve what the user is searching for? Top-ranking pages should score highest.
- "Content Efficiency": Value per word. A focused 2000-word page scoring higher than a bloated 8000-word page is CORRECT.
- "Page Type Fit": Does the page format match what Google rewards for this query?
- "Content Focus vs Bloat": HIGH score = focused, lean, efficient. LOW score = bloated, repetitive, over-structured.
- "Over-Optimization Risk": HIGH score = natural, balanced. LOW score = over-optimized, keyword-stuffed, excessive headings/schema/elements.
- All scores 0-100. A page ranking on page 45 should NOT get the highest scores just because it has more SEO elements.
- NEVER equate "more features" with "better ranking potential".
- Base scores on RANKING REALITY: pages that rank higher should generally score higher on ranking-correlated dimensions."""

    user_prompt = f"""Analyze these pages that rank for the same query. Study what Google is rewarding in the top-ranking pages, then evaluate all pages against those learned preferences.

{page_blocks}

{"The main page (Your Page) needs diagnosis: why is it ranked where it is? Learn from what the top-ranking competitors do differently." if has_main else "Compare these competitor pages to understand what Google prefers for this query."}

Return the complete JSON analysis."""

    return system_prompt, user_prompt


def _build_full_compare_prompt(your_content, your_wc, your_max, fetched, competitor_blocks):
    """Build system/user prompts for user-content-vs-competitors comparison."""
    num_comp = len(fetched)
    comp_word = f"{num_comp} competitor article{'s' if num_comp > 1 else ''}"

    article_entries = [
        f'    {{ "label": "Your Content", "title": "Detected or inferred title", '
        f'"word_count": {your_wc}, "overall_grade": "A+ to F letter grade" }}'
    ]
    for i, f in enumerate(fetched, 1):
        article_entries.append(
            f'    {{ "label": "Competitor {i}", "title": "Detected title", '
            f'"word_count": {f["word_count"]}, "overall_grade": "A+ to F letter grade" }}'
        )
    articles_json = ",\n".join(article_entries)

    scores_csv = "score_yours_0_to_100, " + ", ".join(f"score_comp{i}" for i in range(1, num_comp + 1))
    section_analysis_fields = '"your_analysis": "How your title compares — specific feedback",\n      '
    section_analysis_fields += "\n      ".join(
        f'"competitor_{i}_analysis": "What competitor {i} does",' for i in range(1, num_comp + 1)
    )

    system_prompt = f"""You are an elite content analyst performing a deep, section-by-section and line-by-line comparison between a user's content and {comp_word} that are currently ranking on Google.

Your comparison must be brutally honest, specific, and actionable. Do NOT give generic feedback. Reference specific sentences, paragraphs, and sections.

Return a JSON object with this exact structure:
{{
  "articles": [
{articles_json}
  ],

  "overall_verdict": "A 3-4 sentence executive summary. Where does the user's content stand versus competition? Is it publishable as-is, or does it need work? What is the single biggest gap?",

  "comparative_scores": {{
    "Content Depth & Value": [{scores_csv}],
    "Expertise & Authority Signals": [{scores_csv}],
    "SEO Optimization": [{scores_csv}],
    "Readability & Flow": [{scores_csv}],
    "Originality & Unique Value": [{scores_csv}],
    "Structure & Formatting": [{scores_csv}],
    "Engagement & Hook Quality": [{scores_csv}],
    "Actionability": [{scores_csv}],
    "E-E-A-T & Trust Signals": [{scores_csv}],
    "Page Experience & UX": [{scores_csv}],
    "SERP Listing & CTR Potential": [{scores_csv}]
  }},

  "section_comparisons": [
    {{
      "dimension": "SERP Listing & CTR Potential",
      {section_analysis_fields}
      "your_verdict": "Winning" or "Losing" or "Tied",
      "specific_feedback": "Exactly what to change and why"
    }},
    {{ "dimension": "Title Tag Optimization", ... same structure ... }},
    {{ "dimension": "Meta Description Effectiveness", ... }},
    {{ "dimension": "URL Slug & Structure", ... }},
    {{ "dimension": "Introduction / Hook", ... }},
    {{ "dimension": "Content Depth & Examples", ... }},
    {{ "dimension": "Data & Evidence", ... }},
    {{ "dimension": "Structure & Headings", ... }},
    {{ "dimension": "Writing Quality & Voice", ... }},
    {{ "dimension": "SEO Signals", ... }},
    {{ "dimension": "Visuals & Formatting", ... }},
    {{ "dimension": "Call-to-Action & Conversion", ... }},
    {{ "dimension": "Conclusion & Takeaways", ... }},
    {{ "dimension": "AI vs Human Quality", ... }},
    {{ "dimension": "Unique Value Proposition", ... }},
    {{ "dimension": "E-E-A-T & Trust Signals", ... }},
    {{ "dimension": "Internal Linking Strategy", ... }},
    {{ "dimension": "External Linking & Authority", ... }},
    {{ "dimension": "Page Structure & Schema", ... }},
    {{ "dimension": "Media & Rich Content", ... }}
  ],

  "action_items": [
    "Specific, prioritized action item 1 — what exactly to change/add/remove and why",
    "Action item 2...",
    "... up to 10 action items, ordered by impact"
  ]
}}

IMPORTANT RULES:
- Each article includes a [PAGE CONTEXT] block with SERP LISTING data (title tag with char count, meta description with char count, URL slug, URL depth, Open Graph tags, canonical URL) and PAGE STRUCTURE data (headings, links, images, schema, trust signals). Use ALL of this in your analysis.
- For "SERP Listing & CTR Potential": analyze the title tag (is it under 60 chars? does it contain the primary keyword? is it compelling enough to click?), meta description (is it 120-160 chars? does it include a call-to-action or value prop? does it contain the keyword?), and URL slug (is it clean, keyword-rich, and short?). Compare how each would appear in Google search results.
- For "Title Tag Optimization": deep-dive into title tag construction — keyword placement (front-loaded?), brand inclusion, emotional triggers, power words, uniqueness vs generic phrasing, and pixel-width considerations.
- For "Meta Description Effectiveness": evaluate as a sales pitch for the click — does it promise value? does it differentiate from other SERP results? does it use numbers, questions, or urgency?
- For "URL Slug & Structure": analyze URL readability, keyword inclusion, depth (shorter paths rank better), use of dates (stale signal?), and parameter cleanliness.
- For "Internal Linking Strategy": use the LINKING PROFILE data. Analyze internal link volume (how many?), anchor text quality (are they descriptive and keyword-relevant, or generic like "click here"?), topical clustering (do links point to related content?), and whether the page builds a strong internal link graph. Reference specific anchor text samples from the data.
- For "External Linking & Authority": analyze outbound link quality (are they linking to authoritative sources like research, government, industry leaders?), link volume, nofollow vs dofollow balance, sponsored link disclosure, anchor text descriptiveness, and whether external links add credibility or are just affiliate/commercial. Reference specific external link details and domains from the data.
- Every "your_analysis" and "competitor_X_analysis" must reference SPECIFIC content or context data from the articles, not generic statements.
- For context-focused dimensions (E-E-A-T, Linking, Schema, Media), base your analysis on the [PAGE CONTEXT] metadata, not just the text.
- "your_verdict" must be exactly "Winning", "Losing", or "Tied".
- "specific_feedback" must tell the user EXACTLY what to change — cite the section, the sentence, the issue, and the fix.
- Be harsh and honest. If the user's content is worse, say so clearly.
- All scores are 0-100 integers."""

    user_prompt = f"""Compare these articles in depth. Each article includes both PAGE CONTEXT (SERP listing data + structural metadata) and CONTENT (text). Analyze both layers — pay special attention to how each page would appear in Google search results and what drives clicks.

--- YOUR CONTENT ({your_wc} words) ---
{your_content[:your_max]}
--- END YOUR CONTENT ---
{competitor_blocks}

Perform a deep, section-by-section comparison covering SERP presence, content quality, AND page context/structure. Return the complete JSON analysis."""
    return system_prompt, user_prompt


def _build_competitor_only_prompt(fetched, competitor_blocks, comp_max):
    """Build system/user prompts for competitor-vs-competitor analysis (no user content)."""
    num = len(fetched)
    article_entries = []
    score_entries = []
    for i, f in enumerate(fetched, 1):
        article_entries.append(
            f'    {{ "label": "Competitor {i}", "title": "Detected title", '
            f'"word_count": {f["word_count"]}, "overall_grade": "A+ to F letter grade", '
            f'"url": "{f["url"]}" }}'
        )
        score_entries.append(f"score_comp{i}")
    articles_json = ",\n".join(article_entries)
    scores_csv = ", ".join(score_entries)

    comp_labels = [f"Competitor {i+1}" for i in range(num)]
    section_analysis_fields = "\n      ".join(
        f'"competitor_{i+1}_analysis": "What {comp_labels[i]} does",' for i in range(num)
    )

    system_prompt = f"""You are an elite content analyst performing a deep, head-to-head comparison between {num} competitor articles that are currently ranking on Google on similar topics.

Your analysis must be brutally honest, specific, and actionable. Do NOT give generic feedback. Reference specific sentences, paragraphs, and sections from each article.

Return a JSON object with this exact structure:
{{
  "articles": [
{articles_json}
  ],

  "overall_verdict": "A 3-5 sentence executive summary comparing the competitors head-to-head. Which article is stronger overall? What are the key differentiators? What can be learned from each?",

  "comparative_scores": {{
    "Content Depth & Value": [{scores_csv}],
    "Expertise & Authority Signals": [{scores_csv}],
    "SEO Optimization": [{scores_csv}],
    "Readability & Flow": [{scores_csv}],
    "Originality & Unique Value": [{scores_csv}],
    "Structure & Formatting": [{scores_csv}],
    "Engagement & Hook Quality": [{scores_csv}],
    "Actionability": [{scores_csv}],
    "E-E-A-T & Trust Signals": [{scores_csv}],
    "Page Experience & UX": [{scores_csv}],
    "SERP Listing & CTR Potential": [{scores_csv}]
  }},

  "section_comparisons": [
    {{
      "dimension": "SERP Listing & CTR Potential",
      {section_analysis_fields}
      "winner": "{comp_labels[0]}" or "{comp_labels[1]}" or "Tied",
      "specific_feedback": "Key differences and takeaways"
    }},
    {{ "dimension": "Title Tag Optimization", ... same structure ... }},
    {{ "dimension": "Meta Description Effectiveness", ... }},
    {{ "dimension": "URL Slug & Structure", ... }},
    {{ "dimension": "Introduction / Hook", ... }},
    {{ "dimension": "Content Depth & Examples", ... }},
    {{ "dimension": "Data & Evidence", ... }},
    {{ "dimension": "Structure & Headings", ... }},
    {{ "dimension": "Writing Quality & Voice", ... }},
    {{ "dimension": "SEO Signals", ... }},
    {{ "dimension": "Visuals & Formatting", ... }},
    {{ "dimension": "Call-to-Action & Conversion", ... }},
    {{ "dimension": "Conclusion & Takeaways", ... }},
    {{ "dimension": "AI vs Human Quality", ... }},
    {{ "dimension": "Unique Value Proposition", ... }},
    {{ "dimension": "E-E-A-T & Trust Signals", ... }},
    {{ "dimension": "Internal Linking Strategy", ... }},
    {{ "dimension": "External Linking & Authority", ... }},
    {{ "dimension": "Page Structure & Schema", ... }},
    {{ "dimension": "Media & Rich Content", ... }}
  ],

  "action_items": [
    "Key insight or takeaway 1 — what makes the winning article better in this area",
    "Key insight 2...",
    "... up to 10 insights, ordered by significance"
  ]
}}

IMPORTANT RULES:
- Each article includes a [PAGE CONTEXT] block with SERP LISTING data (title tag with char count, meta description with char count, URL slug, URL depth, Open Graph tags, canonical URL) and PAGE STRUCTURE data (headings, links, images, schema, trust signals). Use ALL of this in your analysis.
- For "SERP Listing & CTR Potential": analyze how each page would appear in Google search results. Which listing would get the most clicks and why?
- For "Title Tag Optimization": deep-dive into title tag construction — keyword placement, length (under 60 chars?), emotional triggers, power words, brand inclusion, uniqueness.
- For "Meta Description Effectiveness": evaluate as a sales pitch for the click — value promise, differentiation, use of numbers/questions/urgency, keyword inclusion, length (120-160 chars optimal).
- For "URL Slug & Structure": analyze readability, keyword inclusion, depth (shorter paths rank better), date presence (stale signal?), and cleanliness.
- For "Internal Linking Strategy": use the LINKING PROFILE data. Analyze internal link volume, anchor text quality (descriptive vs generic), topical clustering, and how well the page builds an internal link graph. Reference specific anchor text samples.
- For "External Linking & Authority": analyze outbound link quality (authoritative sources vs affiliate/commercial?), volume, nofollow vs dofollow balance, sponsored link disclosure, and whether external links add credibility. Reference specific external link details and domains.
- Every "competitor_X_analysis" must reference SPECIFIC content or context data from that article, not generic statements.
- For context-focused dimensions (E-E-A-T, Linking, Schema, Media), base your analysis on the [PAGE CONTEXT] metadata.
- "winner" must be exactly one of the competitor labels or "Tied".
- "specific_feedback" must cite specific sections, sentences, and differences.
- Be harsh and honest. Grade each article on its own merits.
- All scores are 0-100 integers."""

    user_prompt = f"""Analyze and compare these competitor articles in depth. Each article includes both PAGE CONTEXT (SERP listing data + structural metadata) and CONTENT (text). Analyze both layers — pay special attention to how each page would appear in Google search results and what drives clicks.
{competitor_blocks}

Perform a deep, section-by-section comparison covering SERP presence, content quality, AND page context/structure. Return the complete JSON analysis."""
    return system_prompt, user_prompt


_POS_LABELS = {
    "top3": "Top 3 (Position 1-3)",
    "top10": "Page 1 (Position 4-10)",
    "page2": "Page 2 (Position 11-20)",
    "page3_5": "Page 3-5 (Position 21-50)",
    "page5plus": "Page 5+ (Position 50+)",
    "not_ranking": "Not ranking",
    "unknown": "Unknown",
}


@app.post("/api/content-compare")
async def content_compare(request: Request):
    """Ranking-aware comparison: learn what Google prefers, evaluate your page."""
    form = await request.form()
    main_url = form.get("main_url", "").strip()
    main_position = form.get("main_position", "")
    your_content = form.get("your_content", "").strip()
    model = form.get("model", "")
    competitor_urls_json = form.get("competitor_urls", "[]")
    competitor_positions_json = form.get("competitor_positions", "[]")

    try:
        competitor_urls = json.loads(competitor_urls_json)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid competitor_urls JSON.")
    try:
        comp_positions = json.loads(competitor_positions_json)
    except json.JSONDecodeError:
        comp_positions = []

    competitor_urls = [u.strip() for u in competitor_urls if u.strip()]
    while len(comp_positions) < len(competitor_urls):
        comp_positions.append("")
    if not competitor_urls:
        raise HTTPException(status_code=422, detail="Provide at least one competitor URL.")
    if len(competitor_urls) > 10:
        raise HTTPException(status_code=422, detail="Maximum 10 competitor URLs.")

    has_main = bool(main_url)
    has_content = bool(your_content)
    if not has_main and not has_content and len(competitor_urls) < 2:
        raise HTTPException(status_code=422, detail="Without a main URL, provide at least 2 competitor URLs.")

    model_to_use = model if model else DEFAULT_MODEL
    set_model(model_to_use)

    async def safe_fetch(url: str, label: str) -> dict:
        try:
            html = await _fetch_page_html(url)
            text, wc = await _fetch_and_extract(url)
            context = _extract_page_context(html, url)
            return {"url": url, "text": text, "word_count": wc, "label": label,
                    "context": context, "error": None}
        except Exception as e:
            return {"url": url, "text": "", "word_count": 0, "label": label,
                    "context": {}, "error": str(e)}

    all_tasks = []
    if has_main:
        all_tasks.append(safe_fetch(main_url, "Your Page"))
    for i, url in enumerate(competitor_urls):
        all_tasks.append(safe_fetch(url, f"Competitor {i+1}"))

    fetched_all = await asyncio.gather(*all_tasks)

    for f in fetched_all:
        if f["error"]:
            raise HTTPException(status_code=422,
                                detail=f"Failed to fetch {f['label']} ({f['url']}): {f['error']}")

    main_data = fetched_all[0] if has_main else None
    comp_data = fetched_all[1:] if has_main else fetched_all
    num_total = len(fetched_all)
    total_budget = 50000
    per_page = total_budget // max(num_total, 1)

    # Attach position info
    if main_data:
        main_data["position"] = _POS_LABELS.get(main_position, "")
    for i, f in enumerate(comp_data):
        pos = comp_positions[i] if i < len(comp_positions) else ""
        f["position"] = _POS_LABELS.get(pos, "")

    # Build page blocks
    def _page_block(f: dict, max_chars: int) -> str:
        ctx = _format_context_for_prompt(f.get("context", {}))
        pos_line = f"\nRanking Position: {f['position']}" if f.get("position") else ""
        return (
            f"\n\n--- {f['label'].upper()} (URL: {f['url']}, {f['word_count']} words) ---"
            f"{pos_line}"
            f"\n\n[PAGE CONTEXT]\n{ctx}\n[END PAGE CONTEXT]"
            f"\n\n[CONTENT]\n{f['text'][:max_chars]}\n[END CONTENT]"
            f"\n--- END {f['label'].upper()} ---"
        )

    page_blocks = ""
    if main_data:
        page_blocks += _page_block(main_data, per_page)
    for f in comp_data:
        page_blocks += _page_block(f, per_page)

    system_prompt, user_prompt = _build_ranking_analysis_prompt(
        main_data, comp_data, page_blocks, has_main,
    )

    token_budget = 8192 + (num_total * 3000)
    token_budget = min(token_budget, 32768)

    try:
        result = await llm_review(
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            model=model_to_use,
            max_tokens=token_budget,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LLM comparison failed: {e}")

    if result.get("parse_error"):
        raw = result.get("raw_response", "")
        if raw:
            log.warning(f"Comparison JSON parse failed ({len(raw)} chars) — attempting repair")
            repaired = raw.rstrip()
            for _ in range(20):
                if repaired.endswith(","):
                    repaired = repaired[:-1]
                else:
                    break
            opens_b = repaired.count("[") - repaired.count("]")
            opens_c = repaired.count("{") - repaired.count("}")
            repaired += "]" * max(opens_b, 0) + "}" * max(opens_c, 0)
            try:
                result = json.loads(repaired)
                log.info("JSON repair succeeded")
            except json.JSONDecodeError:
                raise HTTPException(
                    status_code=500,
                    detail="LLM response too large to parse. Try fewer competitors.",
                )

    result["model_used"] = model_to_use
    report_id = _save_comparison_report(result)
    result["report_id"] = report_id
    return result


# ---------------------------------------------------------------------------
# Comparison report persistence & downloads
# ---------------------------------------------------------------------------
REPORTS_DIR = Path(__file__).parent / "reports"
REPORTS_DIR.mkdir(exist_ok=True)


def _save_comparison_report(data: dict) -> str:
    """Save comparison result to a JSON file and return the report ID."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_id = f"compare_{ts}"
    path = REPORTS_DIR / f"{report_id}.json"
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    return report_id


def _load_report(report_id: str) -> dict:
    path = REPORTS_DIR / f"{report_id}.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Report not found.")
    return json.loads(path.read_text())


def _report_to_text(data: dict) -> str:
    """Convert comparison JSON to a readable plain-text report."""
    lines = ["=" * 60, "DEEP COMPARISON REPORT", "=" * 60, ""]

    articles = data.get("articles", [])
    for a in articles:
        grade = f"  |  Grade: {a['overall_grade']}" if a.get("overall_grade") else ""
        url_line = f"  |  URL: {a['url']}" if a.get("url") else ""
        lines.append(f"  {a.get('label', '?')}:  {a.get('title', 'Untitled')}  ({a.get('word_count', '?')} words){grade}")
        if url_line:
            lines.append(url_line)
    lines.append("")

    verdict = data.get("overall_verdict", "")
    if verdict:
        lines += ["OVERALL VERDICT", "-" * 40, verdict, ""]

    scores = data.get("comparative_scores", {})
    if scores:
        labels = [a.get("label", f"Article {i+1}") for i, a in enumerate(articles)]
        lines += ["COMPARATIVE SCORES", "-" * 40]
        for dim, vals in scores.items():
            if isinstance(vals, list):
                score_parts = "  |  ".join(
                    f"{labels[i] if i < len(labels) else '?'}: {v}" for i, v in enumerate(vals)
                )
                lines.append(f"  {dim}:  {score_parts}")
        lines.append("")

    sections = data.get("section_comparisons", [])
    if sections:
        lines += ["SECTION-BY-SECTION ANALYSIS", "-" * 40]
        for sec in sections:
            dim = sec.get("dimension", sec.get("section_name", "Section"))
            verdict_label = sec.get("your_verdict", sec.get("winner", ""))
            lines.append(f"\n### {dim}  [{verdict_label}]")
            for key, val in sec.items():
                if key in ("dimension", "section_name", "your_verdict", "winner"):
                    continue
                if val:
                    label = key.replace("_", " ").title()
                    lines.append(f"  {label}: {val}")
        lines.append("")

    action_items = data.get("action_items", [])
    if action_items:
        lines += ["ACTION ITEMS", "-" * 40]
        for i, item in enumerate(action_items, 1):
            lines.append(f"  {i}. {item}")
        lines.append("")

    model = data.get("model_used", "")
    if model:
        lines.append(f"Model: {model}")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 60)

    return "\n".join(lines)


@app.post("/api/comparison-report/download-text")
async def download_comparison_text(request: Request):
    """Generate downloadable plain-text report from comparison data."""
    body = await request.json()
    text = _report_to_text(body)
    return Response(
        content=text,
        media_type="text/plain",
        headers={"Content-Disposition": 'attachment; filename="comparison_report.txt"'},
    )


@app.post("/api/comparison-report/download-excel")
async def download_comparison_excel(request: Request):
    """Generate a polished, client-ready Excel comparison report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    import io

    data = await request.json()
    wb = Workbook()
    articles = data.get("articles", [])
    num_articles = len(articles)

    # ── Style palette ──
    PURPLE = "6D28D9"
    PURPLE_LIGHT = "EDE9FE"
    PURPLE_MID = "A78BFA"
    GREEN = "22C55E"
    GREEN_BG = "F0FDF4"
    BLUE = "3B82F6"
    BLUE_BG = "EFF6FF"
    ORANGE = "F59E0B"
    ORANGE_BG = "FFFBEB"
    RED = "EF4444"
    RED_BG = "FEF2F2"
    DARK = "1E1E1E"
    BODY = "374151"
    MUTED = "6B7280"
    WHITE = "FFFFFF"
    BG_LIGHT = "F9FAFB"
    BG_STRIPE = "F3F4F6"

    header_font = Font(bold=True, size=11, color=WHITE)
    header_fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
    title_font = Font(bold=True, size=14, color=PURPLE)
    subtitle_font = Font(bold=True, size=11, color=PURPLE)
    section_font = Font(bold=True, size=10, color=DARK)
    section_fill = PatternFill(start_color=PURPLE_LIGHT, end_color=PURPLE_LIGHT, fill_type="solid")
    body_font = Font(size=10, color=BODY)
    muted_font = Font(size=9, color=MUTED)
    label_font = Font(bold=True, size=10, color=DARK)
    grade_font = Font(bold=True, size=14, color=PURPLE)
    stripe_fill = PatternFill(start_color=BG_STRIPE, end_color=BG_STRIPE, fill_type="solid")
    thin_side = Side(style="thin", color="D1D5DB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    bottom_border = Border(bottom=Side(style="medium", color=PURPLE))
    wrap_top = Alignment(wrap_text=True, vertical="top")
    wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
    left_top = Alignment(wrap_text=True, vertical="top", horizontal="left")

    def _col(n):
        return get_column_letter(n)

    def _apply_row(ws, row_num, font=None, fill=None, border=None, alignment=None, height=None):
        for cell in ws[row_num]:
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment
        if height:
            ws.row_dimensions[row_num].height = height

    def _score_fill(v):
        v = int(v) if isinstance(v, (int, float)) else 0
        if v >= 70:
            return PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
        if v >= 40:
            return PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid")
        return PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")

    def _score_font(v):
        v = int(v) if isinstance(v, (int, float)) else 0
        if v >= 70:
            return Font(bold=True, size=11, color=GREEN)
        if v >= 40:
            return Font(bold=True, size=11, color=ORANGE)
        return Font(bold=True, size=11, color=RED)

    def _grade_font(g):
        if not g:
            return grade_font
        g = g.upper()
        if g.startswith("A"):
            return Font(bold=True, size=14, color=GREEN)
        if g.startswith("B"):
            return Font(bold=True, size=14, color=BLUE)
        if g.startswith("C"):
            return Font(bold=True, size=14, color=ORANGE)
        return Font(bold=True, size=14, color=RED)

    def _verdict_font(v):
        vl = v.lower() if v else ""
        if "win" in vl:
            return Font(bold=True, size=10, color=GREEN)
        if "los" in vl:
            return Font(bold=True, size=10, color=RED)
        return Font(bold=True, size=10, color=ORANGE)

    def _verdict_fill(v):
        vl = v.lower() if v else ""
        if "win" in vl:
            return PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
        if "los" in vl:
            return PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
        return PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid")

    # ==================================================================
    # SHEET 1: Overview
    # ==================================================================
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = PURPLE

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_articles + 1, 3))
    ws["A1"] = "DEEP COMPARISON REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ts = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    model = data.get("model_used", "N/A")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(num_articles + 1, 3))
    ws["A2"] = f"Generated {ts}  |  Model: {model}"
    ws["A2"].font = muted_font
    ws.row_dimensions[2].height = 18

    # Articles table header (row 4)
    r = 4
    ws.cell(row=r, column=1, value="").font = header_font
    ws.cell(row=r, column=1).fill = header_fill
    ws.cell(row=r, column=1).border = thin_border
    for i, a in enumerate(articles):
        c = ws.cell(row=r, column=i + 2, value=a.get("label", f"Article {i+1}"))
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = wrap_center

    # Article details rows
    field_labels = [
        ("title", "Title"),
        ("word_count", "Word Count"),
        ("overall_grade", "Grade"),
        ("url", "URL"),
    ]
    for fi, (field, label) in enumerate(field_labels):
        r += 1
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=1).fill = section_fill if fi % 2 == 0 else PatternFill()
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=1).alignment = left_top

        for ai, a in enumerate(articles):
            c = ws.cell(row=r, column=ai + 2, value=a.get(field, ""))
            c.border = thin_border
            if fi % 2 == 0:
                c.fill = section_fill
            if field == "overall_grade":
                c.font = _grade_font(a.get(field, ""))
                c.alignment = wrap_center
            elif field == "url":
                c.font = muted_font
                c.alignment = left_top
            elif field == "word_count":
                c.alignment = wrap_center
                c.font = body_font
                c.number_format = "#,##0"
            else:
                c.alignment = left_top
                c.font = body_font
        ws.row_dimensions[r].height = 30 if field == "title" else 22

    # Overall Verdict section
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_articles + 1)
    c = ws.cell(row=r, column=1, value="EXECUTIVE SUMMARY")
    c.font = subtitle_font
    c.fill = section_fill
    c.border = bottom_border
    ws.row_dimensions[r].height = 24

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_articles + 1)
    c = ws.cell(row=r, column=1, value=data.get("overall_verdict", ""))
    c.font = body_font
    c.alignment = left_top
    ws.row_dimensions[r].height = 70

    # Column widths
    ws.column_dimensions["A"].width = 16
    for i in range(num_articles):
        ws.column_dimensions[_col(i + 2)].width = 38

    # ==================================================================
    # SHEET 2: Comparative Scores
    # ==================================================================
    ws2 = wb.create_sheet("Scores")
    ws2.sheet_properties.tabColor = BLUE

    labels = [a.get("label", f"Article {i+1}") for i, a in enumerate(articles)]

    # Title
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_articles + 2)
    ws2["A1"] = "COMPARATIVE SCORES"
    ws2["A1"].font = title_font
    ws2.row_dimensions[1].height = 28

    # Header row
    ws2.cell(row=3, column=1, value="Dimension").font = header_font
    ws2.cell(row=3, column=1).fill = header_fill
    ws2.cell(row=3, column=1).border = thin_border
    for i, lbl in enumerate(labels):
        c = ws2.cell(row=3, column=i + 2, value=lbl)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = wrap_center
    # Winner column
    c = ws2.cell(row=3, column=num_articles + 2, value="Leader")
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border
    c.alignment = wrap_center

    scores = data.get("comparative_scores", {})
    row_num = 3
    for dim, vals in scores.items():
        row_num += 1
        is_stripe = (row_num % 2 == 0)

        ws2.cell(row=row_num, column=1, value=dim).font = section_font
        ws2.cell(row=row_num, column=1).border = thin_border
        ws2.cell(row=row_num, column=1).alignment = left_top
        if is_stripe:
            ws2.cell(row=row_num, column=1).fill = stripe_fill

        if isinstance(vals, list):
            max_val = max(vals) if vals else 0
            for vi, v in enumerate(vals):
                c = ws2.cell(row=row_num, column=vi + 2, value=v)
                c.font = _score_font(v)
                c.fill = _score_fill(v)
                c.border = thin_border
                c.alignment = wrap_center

            # Leader column
            if vals:
                max_idx = vals.index(max_val)
                leader = labels[max_idx] if max_idx < len(labels) else "?"
                c = ws2.cell(row=row_num, column=num_articles + 2, value=leader)
                c.font = Font(bold=True, size=9, color=PURPLE)
                c.border = thin_border
                c.alignment = wrap_center
                if is_stripe:
                    c.fill = stripe_fill

        ws2.row_dimensions[row_num].height = 22

    # Averages row
    if scores:
        row_num += 1
        ws2.cell(row=row_num, column=1, value="AVERAGE").font = Font(bold=True, size=11, color=WHITE)
        ws2.cell(row=row_num, column=1).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
        ws2.cell(row=row_num, column=1).border = thin_border

        all_vals = list(scores.values())
        for ai in range(num_articles):
            article_scores = [v[ai] for v in all_vals if isinstance(v, list) and ai < len(v)]
            avg = round(sum(article_scores) / len(article_scores)) if article_scores else 0
            c = ws2.cell(row=row_num, column=ai + 2, value=avg)
            c.font = Font(bold=True, size=12, color=WHITE)
            c.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
            c.border = thin_border
            c.alignment = wrap_center

        # Overall leader
        avgs = []
        for ai in range(num_articles):
            article_scores = [v[ai] for v in all_vals if isinstance(v, list) and ai < len(v)]
            avgs.append(round(sum(article_scores) / len(article_scores)) if article_scores else 0)
        if avgs:
            best_idx = avgs.index(max(avgs))
            c = ws2.cell(row=row_num, column=num_articles + 2, value=labels[best_idx] if best_idx < len(labels) else "?")
            c.font = Font(bold=True, size=10, color=WHITE)
            c.fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
            c.border = thin_border
            c.alignment = wrap_center

        ws2.row_dimensions[row_num].height = 26

    ws2.column_dimensions["A"].width = 32
    for i in range(num_articles):
        ws2.column_dimensions[_col(i + 2)].width = 18
    ws2.column_dimensions[_col(num_articles + 2)].width = 18

    # ==================================================================
    # SHEET 3: Section-by-Section Analysis
    # ==================================================================
    ws3 = wb.create_sheet("Section Analysis")
    ws3.sheet_properties.tabColor = ORANGE

    # Title
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws3["A1"] = "SECTION-BY-SECTION ANALYSIS"
    ws3["A1"].font = title_font
    ws3.row_dimensions[1].height = 28

    sections = data.get("section_comparisons", [])

    # Build structured columns: Dimension | analysis per article | Verdict | Feedback
    analysis_col_keys = []
    for sec in sections:
        for k in sec.keys():
            if k.endswith("_analysis") and k not in analysis_col_keys:
                analysis_col_keys.append(k)

    col_headers = ["Section"]
    col_headers += [k.replace("_", " ").replace("analysis", "").strip().title() for k in analysis_col_keys]
    col_headers += ["Verdict", "Specific Feedback"]

    # Header row
    for ci, h in enumerate(col_headers):
        c = ws3.cell(row=3, column=ci + 1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = wrap_center
    ws3.row_dimensions[3].height = 22

    for si, sec in enumerate(sections):
        r = si + 4
        dim = sec.get("dimension", sec.get("section_name", f"Section {si + 1}"))
        verdict = sec.get("your_verdict", sec.get("winner", ""))
        feedback = sec.get("specific_feedback", "")
        is_stripe = si % 2 == 1

        # Dimension
        c = ws3.cell(row=r, column=1, value=dim)
        c.font = section_font
        c.border = thin_border
        c.alignment = left_top
        if is_stripe:
            c.fill = stripe_fill

        # Analysis columns
        for ki, k in enumerate(analysis_col_keys):
            val = sec.get(k, "")
            c = ws3.cell(row=r, column=ki + 2, value=val)
            c.font = body_font
            c.border = thin_border
            c.alignment = left_top
            if is_stripe:
                c.fill = stripe_fill

        # Verdict
        vc = len(analysis_col_keys) + 2
        c = ws3.cell(row=r, column=vc, value=verdict)
        c.font = _verdict_font(verdict)
        c.fill = _verdict_fill(verdict)
        c.border = thin_border
        c.alignment = wrap_center

        # Specific feedback
        fc = vc + 1
        c = ws3.cell(row=r, column=fc, value=feedback)
        c.font = body_font
        c.border = thin_border
        c.alignment = left_top
        if feedback:
            c.fill = PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid")

        ws3.row_dimensions[r].height = 80

    # Column widths
    ws3.column_dimensions["A"].width = 22
    for ki in range(len(analysis_col_keys)):
        ws3.column_dimensions[_col(ki + 2)].width = 42
    ws3.column_dimensions[_col(len(analysis_col_keys) + 2)].width = 14
    ws3.column_dimensions[_col(len(analysis_col_keys) + 3)].width = 50

    # ==================================================================
    # SHEET 4: Action Items
    # ==================================================================
    ws4 = wb.create_sheet("Action Items")
    ws4.sheet_properties.tabColor = GREEN

    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws4["A1"] = "ACTION ITEMS"
    ws4["A1"].font = title_font
    ws4.row_dimensions[1].height = 28

    ws4.cell(row=3, column=1, value="#").font = header_font
    ws4.cell(row=3, column=1).fill = header_fill
    ws4.cell(row=3, column=1).border = thin_border
    ws4.cell(row=3, column=1).alignment = wrap_center
    ws4.cell(row=3, column=2, value="Priority Action Item").font = header_font
    ws4.cell(row=3, column=2).fill = header_fill
    ws4.cell(row=3, column=2).border = thin_border
    ws4.cell(row=3, column=3, value="Status").font = header_font
    ws4.cell(row=3, column=3).fill = header_fill
    ws4.cell(row=3, column=3).border = thin_border
    ws4.cell(row=3, column=3).alignment = wrap_center

    for i, item in enumerate(data.get("action_items", []), 1):
        r = i + 3
        is_stripe = i % 2 == 0

        c = ws4.cell(row=r, column=1, value=i)
        c.font = Font(bold=True, size=11, color=PURPLE)
        c.border = thin_border
        c.alignment = wrap_center
        if is_stripe:
            c.fill = stripe_fill

        c = ws4.cell(row=r, column=2, value=item)
        c.font = body_font
        c.border = thin_border
        c.alignment = left_top
        if is_stripe:
            c.fill = stripe_fill

        c = ws4.cell(row=r, column=3, value="Pending")
        c.font = Font(size=9, color=MUTED)
        c.border = thin_border
        c.alignment = wrap_center
        if is_stripe:
            c.fill = stripe_fill

        ws4.row_dimensions[r].height = 35

    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 85
    ws4.column_dimensions["C"].width = 12

    # ==================================================================
    # Freeze panes for usability
    # ==================================================================
    ws.freeze_panes = "A5"
    ws2.freeze_panes = "A4"
    ws3.freeze_panes = "A4"
    ws4.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="comparison_report.xlsx"'},
    )


def _sanitize_for_pdf(text: str) -> str:
    """Replace Unicode characters that crash fpdf2's core fonts."""
    replacements = {
        "\u2014": "--", "\u2013": "-", "\u2012": "-",
        "\u2018": "'", "\u2019": "'", "\u201a": "'",
        "\u201c": '"', "\u201d": '"', "\u201e": '"',
        "\u2026": "...", "\u2022": "*", "\u2023": ">",
        "\u2027": "-", "\u2043": "-", "\u25aa": "*",
        "\u25b6": ">", "\u25cf": "*", "\u2192": "->",
        "\u2190": "<-", "\u2194": "<->", "\u2713": "[x]",
        "\u2714": "[x]", "\u2717": "[ ]", "\u2718": "[ ]",
        "\u00a0": " ", "\u200b": "", "\u200c": "", "\u200d": "",
        "\ufeff": "", "\u2028": "\n", "\u2029": "\n",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.encode("latin-1", errors="replace").decode("latin-1")


@app.post("/api/comparison-report/download-pdf")
async def download_comparison_pdf(request: Request):
    """Generate a polished, client-ready PDF comparison report."""
    from fpdf import FPDF
    import io

    data = await request.json()
    S = _sanitize_for_pdf

    PURPLE = (109, 40, 217)
    PURPLE_LIGHT = (237, 233, 254)
    DARK = (30, 30, 30)
    BODY = (55, 55, 55)
    MUTED = (120, 120, 120)
    WHITE = (255, 255, 255)
    GREEN = (34, 197, 94)
    ORANGE = (245, 158, 11)
    RED = (239, 68, 68)
    BLUE = (59, 130, 246)
    BG_LIGHT = (248, 247, 252)
    BG_STRIPE = (243, 244, 246)
    BORDER = (209, 213, 219)
    AMBER_BG = (255, 251, 235)

    def _sc(v):
        v = int(v) if isinstance(v, (int, float)) else 0
        return GREEN if v >= 70 else (ORANGE if v >= 40 else RED)

    def _gc(g):
        if not g: return BODY
        g = g.upper()
        if g.startswith("A"): return GREEN
        if g.startswith("B"): return BLUE
        if g.startswith("C"): return ORANGE
        return RED

    class PDF(FPDF):
        def header(self):
            if self.page_no() == 1:
                return
            self.set_font("Helvetica", "", 7)
            self.set_text_color(*MUTED)
            self.set_y(8)
            self.cell(0, 5, "Vizup Soul  |  Deep Comparison Report", align="R")
            self.set_draw_color(*PURPLE)
            self.line(10, 14, 200, 14)
            self.set_y(17)
        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "", 7)
            self.set_text_color(*MUTED)
            self.cell(0, 5, f"Page {self.page_no()}/{{nb}}", align="C")

    pdf = PDF(orientation="P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=16)
    LM = pdf.l_margin
    PW = pdf.w - pdf.l_margin - pdf.r_margin

    articles = data.get("articles", [])
    n_art = len(articles)
    labels = [S(a.get("label", f"Article {i+1}")) for i, a in enumerate(articles)]

    # ── Helpers ──
    def section_heading(title):
        if pdf.get_y() > pdf.h - 35:
            pdf.add_page()
        pdf.ln(3)
        pdf.set_fill_color(*PURPLE)
        pdf.rect(LM, pdf.get_y(), PW, 9, "F")
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(*WHITE)
        pdf.set_x(LM + 4)
        pdf.cell(PW - 8, 9, S(title), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    def table_row(cells, widths, bold=False, fills=None, fonts=None, h=7, aligns=None):
        y0 = pdf.get_y()
        max_h = h
        rendered = []
        for ci, (txt, w) in enumerate(zip(cells, widths)):
            x = LM + sum(widths[:ci])
            pdf.set_xy(x, y0)
            f = fonts[ci] if fonts else None
            if f:
                pdf.set_font(*f)
            elif bold:
                pdf.set_font("Helvetica", "B", 8)
            else:
                pdf.set_font("Helvetica", "", 8)
            al = (aligns[ci] if aligns else "L")
            pdf.set_text_color(*BODY)
            if fills and fills[ci]:
                pdf.set_fill_color(*fills[ci])
                pdf.multi_cell(w, 4.5, S(str(txt)), align=al, fill=True)
            else:
                pdf.multi_cell(w, 4.5, S(str(txt)), align=al)
            cell_h = pdf.get_y() - y0
            if cell_h > max_h:
                max_h = cell_h
            rendered.append((x, w))
        # Draw cell borders
        pdf.set_draw_color(*BORDER)
        for x, w in rendered:
            pdf.rect(x, y0, w, max_h)
        pdf.set_y(y0 + max_h)

    # ================================================================
    # PAGE 1 — COVER
    # ================================================================
    pdf.add_page()
    pdf.set_fill_color(*PURPLE)
    pdf.rect(0, 0, 210, 48, "F")
    pdf.set_y(10)
    pdf.set_font("Helvetica", "B", 24)
    pdf.set_text_color(*WHITE)
    pdf.cell(0, 12, "Deep Comparison Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(220, 210, 255)
    pdf.cell(0, 6, "Vizup Soul  -  Content Quality Firewall", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    ts = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    pdf.cell(0, 5, S(f"Generated {ts}  |  Model: {data.get('model_used', 'N/A')}"), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_y(55)

    # ── Articles overview as a table (works for any number) ──
    if articles:
        section_heading("Articles Compared")
        col1_w = 28
        grade_w = 16
        rest_w = PW - col1_w - grade_w
        # Header
        pdf.set_fill_color(*PURPLE)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(*WHITE)
        y0 = pdf.get_y()
        for ci, (txt, w) in enumerate([("Article", col1_w), ("Details", rest_w), ("Grade", grade_w)]):
            pdf.set_xy(LM + sum([col1_w, rest_w, grade_w][:ci]), y0)
            pdf.cell(w, 7, txt, align="C", fill=True)
        pdf.set_draw_color(*BORDER)
        pdf.rect(LM, y0, PW, 7)
        pdf.set_y(y0 + 7)

        for i, a in enumerate(articles):
            if pdf.get_y() > pdf.h - 25:
                pdf.add_page()
            y0 = pdf.get_y()
            bg = BG_STRIPE if i % 2 == 1 else None

            # Col 1: Label
            pdf.set_xy(LM, y0)
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(*PURPLE)
            if bg:
                pdf.set_fill_color(*bg)
                pdf.multi_cell(col1_w, 4.5, S(a.get("label", f"Art. {i+1}")), fill=True)
            else:
                pdf.multi_cell(col1_w, 4.5, S(a.get("label", f"Art. {i+1}")))

            # Col 2: Title + word count + URL
            pdf.set_xy(LM + col1_w, y0)
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(*DARK)
            title_str = S(a.get("title", "Untitled"))
            wc = a.get("word_count", "?")
            url = a.get("url", "")
            detail = f"{title_str}\n{wc} words"
            if url:
                url_short = url if len(url) <= 70 else url[:67] + "..."
                detail += f"\n{url_short}"
            if bg:
                pdf.set_fill_color(*bg)
                pdf.multi_cell(rest_w, 4.5, detail, fill=True)
            else:
                pdf.multi_cell(rest_w, 4.5, detail)

            # Col 3: Grade badge
            grade = a.get("overall_grade", "")
            cell_h = max(pdf.get_y() - y0, 14)
            gx = LM + col1_w + rest_w
            if bg:
                pdf.set_fill_color(*bg)
                pdf.rect(gx, y0, grade_w, cell_h, "F")
            if grade:
                gy = y0 + (cell_h - 8) / 2
                gc = _gc(grade)
                pdf.set_fill_color(*gc)
                pdf.rect(gx + 2, gy, grade_w - 4, 8, "F")
                pdf.set_xy(gx + 2, gy + 1)
                pdf.set_font("Helvetica", "B", 11)
                pdf.set_text_color(*WHITE)
                pdf.cell(grade_w - 4, 6, S(grade), align="C")

            # Row border
            pdf.set_draw_color(*BORDER)
            pdf.rect(LM, y0, col1_w, cell_h)
            pdf.rect(LM + col1_w, y0, rest_w, cell_h)
            pdf.rect(gx, y0, grade_w, cell_h)
            pdf.set_y(y0 + cell_h)

    # ── Executive Summary ──
    verdict = data.get("overall_verdict", "")
    if verdict:
        pdf.ln(4)
        section_heading("Executive Summary")
        y0 = pdf.get_y()
        pdf.set_fill_color(*PURPLE_LIGHT)
        pdf.set_x(LM + 4)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*BODY)
        pdf.multi_cell(PW - 8, 5, S(verdict))
        ya = pdf.get_y()
        # Background fill behind text
        pdf.set_fill_color(*PURPLE)
        pdf.rect(LM, y0, 2.5, ya - y0, "F")
        pdf.set_y(ya + 2)

    # ================================================================
    # COMPARATIVE SCORES — clean data table
    # ================================================================
    scores = data.get("comparative_scores", {})
    if scores:
        pdf.add_page()
        section_heading("Comparative Scores")

        # Calculate column widths
        dim_w = 50
        score_col_w = min(28, (PW - dim_w - 22) / max(n_art, 1))
        leader_w = PW - dim_w - score_col_w * n_art

        # Table header
        hdrs = ["Dimension"] + [l[:14] for l in labels] + ["Leader"]
        hdr_ws = [dim_w] + [score_col_w] * n_art + [leader_w]
        y0 = pdf.get_y()
        pdf.set_fill_color(*PURPLE)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(*WHITE)
        for ci, (txt, w) in enumerate(zip(hdrs, hdr_ws)):
            pdf.set_xy(LM + sum(hdr_ws[:ci]), y0)
            pdf.cell(w, 7, S(txt), align="C", fill=True)
        pdf.set_draw_color(*BORDER)
        pdf.rect(LM, y0, PW, 7)
        pdf.set_y(y0 + 7)

        # Score rows
        row_idx = 0
        for dim, vals in scores.items():
            if not isinstance(vals, list):
                continue
            if pdf.get_y() > pdf.h - 18:
                pdf.add_page()
            y0 = pdf.get_y()
            is_stripe = row_idx % 2 == 1
            row_idx += 1

            # Dimension name
            pdf.set_xy(LM, y0)
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(*DARK)
            if is_stripe:
                pdf.set_fill_color(*BG_STRIPE)
                pdf.cell(dim_w, 8, S(dim), fill=True)
            else:
                pdf.cell(dim_w, 8, S(dim))

            # Score cells with color + bar
            max_v = max(vals) if vals else 0
            for vi, v in enumerate(vals):
                score = int(v) if isinstance(v, (int, float)) else 0
                cx = LM + dim_w + vi * score_col_w
                sc = _sc(score)

                # Cell background
                if is_stripe:
                    pdf.set_fill_color(*BG_STRIPE)
                    pdf.rect(cx, y0, score_col_w, 8, "F")

                # Mini bar (bottom of cell)
                bar_w = max(1, (score_col_w - 6) * score / 100)
                pdf.set_fill_color(*sc)
                pdf.rect(cx + 3, y0 + 6, bar_w, 1.5, "F")

                # Score number
                pdf.set_xy(cx, y0)
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(*sc)
                pdf.cell(score_col_w, 6, str(score), align="C")

            # Leader cell
            lx = LM + dim_w + n_art * score_col_w
            best_idx = vals.index(max_v) if vals else 0
            leader_name = labels[best_idx] if best_idx < len(labels) else "?"
            if is_stripe:
                pdf.set_fill_color(*BG_STRIPE)
                pdf.rect(lx, y0, leader_w, 8, "F")
            pdf.set_xy(lx, y0)
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_text_color(*PURPLE)
            pdf.cell(leader_w, 8, S(leader_name), align="C")

            # Row borders
            pdf.set_draw_color(*BORDER)
            x_off = LM
            for w in hdr_ws:
                pdf.rect(x_off, y0, w, 8)
                x_off += w
            pdf.set_y(y0 + 8)

        # Averages row
        y0 = pdf.get_y()
        pdf.set_fill_color(*DARK)
        pdf.rect(LM, y0, PW, 9, "F")
        pdf.set_xy(LM, y0)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(*WHITE)
        pdf.cell(dim_w, 9, "  AVERAGE")

        all_vals = [v for v in scores.values() if isinstance(v, list)]
        avgs = []
        for ai in range(n_art):
            sc_list = [v[ai] for v in all_vals if ai < len(v)]
            avg = round(sum(sc_list) / len(sc_list)) if sc_list else 0
            avgs.append(avg)
            pdf.set_xy(LM + dim_w + ai * score_col_w, y0)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(score_col_w, 9, str(avg), align="C")

        if avgs:
            best_avg_idx = avgs.index(max(avgs))
            pdf.set_xy(LM + dim_w + n_art * score_col_w, y0)
            pdf.set_fill_color(*PURPLE)
            pdf.rect(LM + dim_w + n_art * score_col_w, y0, leader_w, 9, "F")
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(leader_w, 9, S(labels[best_avg_idx] if best_avg_idx < len(labels) else "?"), align="C")
        pdf.set_y(y0 + 9)

    # ================================================================
    # SECTION-BY-SECTION ANALYSIS
    # ================================================================
    sections = data.get("section_comparisons", [])
    if sections:
        pdf.add_page()
        section_heading("Section-by-Section Analysis")

        for si, sec in enumerate(sections):
            dim = S(sec.get("dimension", sec.get("section_name", f"Section {si+1}")))
            verdict_val = sec.get("your_verdict", sec.get("winner", ""))

            if pdf.get_y() > pdf.h - 45:
                pdf.add_page()

            # Section title bar
            y0 = pdf.get_y()
            pdf.set_fill_color(*PURPLE_LIGHT)
            pdf.rect(LM, y0, PW, 8, "F")
            pdf.set_fill_color(*PURPLE)
            pdf.rect(LM, y0, 3, 8, "F")
            pdf.set_xy(LM + 5, y0)
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(*PURPLE)
            pdf.cell(PW - 40, 8, dim)

            # Verdict badge on the right
            if verdict_val:
                vl = verdict_val.lower()
                vc = GREEN if "win" in vl else (RED if "los" in vl else ORANGE)
                vtext = S(verdict_val)
                vw = pdf.get_string_width(vtext) + 8
                vx = LM + PW - vw - 3
                pdf.set_fill_color(*vc)
                pdf.rect(vx, y0 + 1.5, vw, 5, "F")
                pdf.set_xy(vx, y0 + 1.5)
                pdf.set_font("Helvetica", "B", 6)
                pdf.set_text_color(*WHITE)
                pdf.cell(vw, 5, vtext, align="C")
            pdf.set_y(y0 + 9)

            # Analysis entries
            analysis_keys = [k for k in sec.keys() if k.endswith("_analysis") and sec.get(k)]
            feedback = sec.get("specific_feedback", "")

            for ki, key in enumerate(analysis_keys):
                val = sec[key]
                label = key.replace("_analysis", "").replace("_", " ").strip().title()
                if pdf.get_y() > pdf.h - 20:
                    pdf.add_page()

                # Label
                pdf.set_font("Helvetica", "B", 7)
                pdf.set_text_color(*PURPLE)
                pdf.set_x(LM + 3)
                pdf.cell(PW, 5, S(label), new_x="LMARGIN", new_y="NEXT")

                # Analysis text
                pdf.set_font("Helvetica", "", 8)
                pdf.set_text_color(*BODY)
                pdf.set_x(LM + 3)
                pdf.multi_cell(PW - 6, 4, S(str(val)))
                pdf.ln(1)

            # Specific feedback callout
            if feedback:
                if pdf.get_y() > pdf.h - 20:
                    pdf.add_page()
                y0 = pdf.get_y()
                # Draw amber background and left bar first
                # Estimate height: ~4.5 chars/mm at font 8, plus header
                feedback_s = S(feedback)
                est_lines = max(1, len(feedback_s) / ((PW - 6) / 1.8))
                est_h = est_lines * 4 + 6
                pdf.set_fill_color(*AMBER_BG)
                pdf.rect(LM, y0, PW, est_h, "F")
                pdf.set_fill_color(*ORANGE)
                pdf.rect(LM, y0, 2, est_h, "F")
                # Render label + text on top
                pdf.set_xy(LM + 5, y0 + 1)
                pdf.set_font("Helvetica", "B", 7)
                pdf.set_text_color(*ORANGE)
                pdf.cell(PW - 8, 4, "KEY TAKEAWAY", new_x="LMARGIN", new_y="NEXT")
                pdf.set_x(LM + 5)
                pdf.set_font("Helvetica", "", 8)
                pdf.set_text_color(*BODY)
                pdf.multi_cell(PW - 10, 4, feedback_s)
                ya = pdf.get_y()
                # Correct background height if estimate was wrong
                if ya - y0 > est_h:
                    pdf.set_fill_color(*AMBER_BG)
                    pdf.rect(LM, y0 + est_h, PW, ya - y0 - est_h, "F")
                    pdf.set_fill_color(*ORANGE)
                    pdf.rect(LM, y0 + est_h, 2, ya - y0 - est_h, "F")
                pdf.ln(1)

            # Divider
            pdf.set_draw_color(*BORDER)
            pdf.line(LM + 10, pdf.get_y(), LM + PW - 10, pdf.get_y())
            pdf.ln(3)

    # ================================================================
    # ACTION ITEMS
    # ================================================================
    action_items = data.get("action_items", [])
    if action_items:
        if pdf.get_y() > pdf.h - 50:
            pdf.add_page()
        section_heading("Action Items")

        for i, item in enumerate(action_items, 1):
            if pdf.get_y() > pdf.h - 18:
                pdf.add_page()

            y0 = pdf.get_y()
            item_s = S(item)

            # Pre-draw alternating background (overestimate height)
            if i % 2 == 0:
                est_h = max(10, len(item_s) / ((PW - 16) / 1.8) * 4.5 + 4)
                pdf.set_fill_color(*BG_LIGHT)
                pdf.rect(LM, y0, PW, est_h, "F")

            # Number badge
            pdf.set_fill_color(*PURPLE)
            pdf.rect(LM + 2, y0 + 1, 8, 6, "F")
            pdf.set_xy(LM + 2, y0 + 1)
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(*WHITE)
            pdf.cell(8, 6, str(i), align="C")

            # Item text
            pdf.set_xy(LM + 14, y0 + 1)
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(*BODY)
            pdf.multi_cell(PW - 16, 4.5, item_s)
            ya = pdf.get_y()

            # Extend background if needed
            if i % 2 == 0 and ya - y0 > est_h:
                pdf.set_fill_color(*BG_LIGHT)
                pdf.rect(LM, y0 + est_h, PW, ya - y0 - est_h, "F")

            pdf.set_y(ya + 3)

    # ── Output ──
    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="comparison_report.pdf"'},
    )


# ---------------------------------------------------------------------------
# Content Gap Analyzer — Sitemap Discovery, Crawling & Gap Analysis
# ---------------------------------------------------------------------------

_SITEMAP_NS = "{http://www.sitemaps.org/schemas/sitemap/0.9}"
_GAP_MAX_URLS = 500
_GAP_METADATA_CONCURRENCY = 10

_CMS_SITEMAP_PATHS: dict[str, list[str]] = {
    "wordpress_core": ["/wp-sitemap.xml"],
    "wordpress_yoast": ["/sitemap_index.xml", "/post-sitemap.xml", "/page-sitemap.xml"],
    "wordpress_rankmath": ["/sitemap_index.xml"],
    "wordpress_aioseo": ["/sitemap.xml"],
    "shopify": ["/sitemap.xml"],
    "ghost": ["/sitemap.xml", "/sitemap-posts.xml", "/sitemap-pages.xml"],
    "general": ["/sitemap.xml", "/sitemap_index.xml", "/sitemap.xml.gz",
                "/post-sitemap.xml", "/page-sitemap.xml", "/blog-sitemap.xml",
                "/wp-sitemap.xml"],
}

_BLOG_URL_RE = re.compile(
    r"/(blog|news|magazine|articles?|insights?|resources|journal|stories|posts?|updates?|press)"
    r"|/\d{4}/\d{2}/",
    re.IGNORECASE,
)
_LANDING_URL_RE = re.compile(
    r"/(services?|products?|features?|solutions?|pricing|about|contact|platform|"
    r"integrations?|demo|trial|signup|enterprise)/?$",
    re.IGNORECASE,
)
_SKIP_URL_RE = re.compile(
    r"/(tag|category|categories|author|authors|page/\d|feed|rss|amp|print|embed|"
    r"attachment|wp-content|wp-admin|cart|checkout|account|login|search)"
    r"|#|\.(pdf|jpg|jpeg|png|gif|svg|css|js|xml|zip|mp4|mp3)$",
    re.IGNORECASE,
)


def _normalize_domain(raw: str) -> str:
    raw = raw.strip().rstrip("/")
    if not raw.startswith("http"):
        raw = f"https://{raw}"
    return raw


async def _gap_fetch(url: str, timeout: float = 15.0,
                     accept: str = "*/*") -> httpx.Response | None:
    try:
        async with httpx.AsyncClient(
            follow_redirects=True, timeout=timeout,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/131.0.0.0 Safari/537.36"
                ),
                "Accept": accept,
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
            },
        ) as client:
            resp = await client.get(url)
            if resp.status_code < 400:
                return resp
            resp._gap_error = f"HTTP {resp.status_code}"
            return resp
    except Exception as e:
        return None


def _gap_resp_ok(resp: httpx.Response | None) -> bool:
    """Check if a gap_fetch response is usable (not None, not error status)."""
    return resp is not None and resp.status_code < 400


def _detect_cms(html: str) -> str | None:
    hl = html.lower()
    if "wp-content" in hl or "wp-includes" in hl or 'name="generator" content="wordpress' in hl:
        if "yoast" in hl:
            return "wordpress_yoast"
        if "rank math" in hl or "rank-math" in hl:
            return "wordpress_rankmath"
        if "aioseo" in hl or "all in one seo" in hl:
            return "wordpress_aioseo"
        return "wordpress_core"
    if "shopify" in hl or "cdn.shopify" in hl:
        return "shopify"
    if '"ghost"' in hl or "ghost.io" in hl or 'name="generator" content="ghost' in hl:
        return "ghost"
    if "squarespace" in hl:
        return "squarespace"
    if "wix.com" in hl:
        return "wix"
    if "webflow" in hl:
        return "webflow"
    return None


async def _discover_sitemaps(domain: str) -> dict:
    base = _normalize_domain(domain)
    parsed = urlparse(base)
    clean_domain = parsed.netloc or parsed.path.split("/")[0]

    result: dict[str, Any] = {
        "domain": clean_domain,
        "base_url": base,
        "discovery_method": [],
        "cms_detected": None,
        "sitemap_urls_found": [],
        "errors": [],
    }
    found_sitemaps: set[str] = set()

    # Strategy 1: robots.txt
    resp = await _gap_fetch(f"{base}/robots.txt")
    if resp and resp.status_code == 200:
        for line in resp.text.splitlines():
            if line.strip().lower().startswith("sitemap:"):
                sm_url = line.split(":", 1)[1].strip()
                if sm_url:
                    found_sitemaps.add(sm_url)
                    if "robots.txt" not in result["discovery_method"]:
                        result["discovery_method"].append("robots.txt")

    # Strategy 2: CMS-aware probing
    homepage_resp = await _gap_fetch(base, accept="text/html")
    if homepage_resp and homepage_resp.status_code == 200:
        cms = _detect_cms(homepage_resp.text)
        result["cms_detected"] = cms
        if cms and cms in _CMS_SITEMAP_PATHS:
            for path in _CMS_SITEMAP_PATHS[cms]:
                sm_url = f"{base}{path}"
                if sm_url not in found_sitemaps:
                    r2 = await _gap_fetch(sm_url, accept="application/xml, text/xml, */*")
                    if r2 and ("<urlset" in r2.text or "<sitemapindex" in r2.text):
                        found_sitemaps.add(sm_url)
                        if "CMS probe" not in " ".join(result["discovery_method"]):
                            result["discovery_method"].append(f"CMS probe ({cms})")

    # Strategy 3: Fallback common paths
    if not found_sitemaps:
        for path in _CMS_SITEMAP_PATHS["general"]:
            sm_url = f"{base}{path}"
            r2 = await _gap_fetch(sm_url, accept="application/xml, text/xml, */*")
            if r2 and r2.status_code == 200:
                text = r2.text
                if path.endswith(".gz"):
                    try:
                        text = gzip.decompress(r2.content).decode("utf-8")
                    except Exception:
                        continue
                if "<urlset" in text or "<sitemapindex" in text:
                    found_sitemaps.add(sm_url)
                    if "fallback probe" not in result["discovery_method"]:
                        result["discovery_method"].append("fallback probe")
                    break

    result["sitemap_urls_found"] = list(found_sitemaps)
    if not found_sitemaps:
        result["errors"].append("No sitemaps found")
    return result


async def _parse_sitemap(
    sitemap_url: str, max_depth: int = 3,
    _seen: set | None = None, _depth: int = 0,
) -> dict:
    if _seen is None:
        _seen = set()
    if sitemap_url in _seen or _depth > max_depth:
        return {"sitemap_tree": [], "urls": [], "errors": []}

    _seen.add(sitemap_url)
    result: dict[str, Any] = {"sitemap_tree": [], "urls": [], "errors": []}

    resp = await _gap_fetch(sitemap_url, accept="application/xml, text/xml, */*")
    if not resp:
        result["errors"].append(f"Connection failed: {sitemap_url}")
        return result
    if resp.status_code >= 400:
        result["errors"].append(f"HTTP {resp.status_code} (not found): {sitemap_url}")
        return result

    xml_text = resp.text
    if sitemap_url.endswith(".gz"):
        try:
            xml_text = gzip.decompress(resp.content).decode("utf-8")
        except Exception as e:
            result["errors"].append(f"Gzip error: {sitemap_url}: {e}")
            return result

    if "<urlset" not in xml_text and "<sitemapindex" not in xml_text:
        result["errors"].append(f"Not a valid sitemap (no XML urlset/index): {sitemap_url}")
        return result

    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as e:
        result["errors"].append(f"XML parse error: {sitemap_url}: {e}")
        return result

    root_tag = root.tag.split("}")[-1] if "}" in root.tag else root.tag

    if root_tag == "sitemapindex":
        children = []
        for sm_el in root.findall(f"{_SITEMAP_NS}sitemap"):
            loc_el = sm_el.find(f"{_SITEMAP_NS}loc")
            if loc_el is not None and loc_el.text:
                child_url = loc_el.text.strip()
                child = await _parse_sitemap(child_url, max_depth, _seen, _depth + 1)
                children.append({
                    "url": child_url, "type": "child_sitemap",
                    "url_count": len(child["urls"]),
                })
                result["urls"].extend(child["urls"])
                result["errors"].extend(child["errors"])
        result["sitemap_tree"].append({
            "url": sitemap_url, "type": "index", "children": children,
        })

    elif root_tag == "urlset":
        extracted = []
        for url_el in root.findall(f"{_SITEMAP_NS}url"):
            loc_el = url_el.find(f"{_SITEMAP_NS}loc")
            if loc_el is None or not loc_el.text:
                continue
            entry: dict[str, str] = {"url": loc_el.text.strip()}
            lm_el = url_el.find(f"{_SITEMAP_NS}lastmod")
            if lm_el is not None and lm_el.text:
                entry["lastmod"] = lm_el.text.strip()
            cf_el = url_el.find(f"{_SITEMAP_NS}changefreq")
            if cf_el is not None and cf_el.text:
                entry["changefreq"] = cf_el.text.strip()
            pr_el = url_el.find(f"{_SITEMAP_NS}priority")
            if pr_el is not None and pr_el.text:
                entry["priority"] = pr_el.text.strip()
            extracted.append(entry)
        result["urls"] = extracted
        result["sitemap_tree"].append({
            "url": sitemap_url, "type": "urlset", "url_count": len(extracted),
        })
    else:
        result["errors"].append(f"Unknown root element: {root_tag} in {sitemap_url}")

    return result


def _classify_gap_urls(urls: list[dict], domain: str) -> dict:
    parsed_domain = urlparse(_normalize_domain(domain)).netloc
    blog_posts, landing_pages, other = [], [], []

    for entry in urls:
        url = entry.get("url", "")
        parsed = urlparse(url)
        if parsed.netloc and parsed.netloc != parsed_domain and \
                not parsed.netloc.endswith(f".{parsed_domain}"):
            continue
        path = parsed.path.rstrip("/")
        if _SKIP_URL_RE.search(path):
            other.append(entry)
            continue
        if not path or path == "/":
            landing_pages.append(entry)
            continue
        segments = [s for s in path.strip("/").split("/") if s]
        depth = len(segments)
        if _BLOG_URL_RE.search(path):
            blog_posts.append(entry)
        elif _LANDING_URL_RE.search(path) or depth <= 2:
            landing_pages.append(entry)
        elif depth >= 3:
            blog_posts.append(entry)
        else:
            other.append(entry)

    return {
        "blog_posts": blog_posts, "landing_pages": landing_pages,
        "other": other,
        "total": len(blog_posts) + len(landing_pages) + len(other),
    }


def _compute_freshness(urls: list[dict]) -> dict:
    now = datetime.now()
    dates: list[datetime] = []
    for entry in urls:
        lm = entry.get("lastmod", "")
        if not lm:
            continue
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%SZ"):
            try:
                dates.append(datetime.strptime(lm[:19], fmt[:19]))
                break
            except ValueError:
                continue

    if not dates:
        return {"oldest_lastmod": None, "newest_lastmod": None,
                "pct_with_lastmod": 0,
                "last_30d": 0, "last_90d": 0, "last_180d": 0, "last_365d": 0}

    dates.sort()
    total = max(len(urls), 1)
    return {
        "oldest_lastmod": dates[0].strftime("%Y-%m-%d"),
        "newest_lastmod": dates[-1].strftime("%Y-%m-%d"),
        "pct_with_lastmod": round(len(dates) / total * 100),
        "last_30d": sum(1 for d in dates if (now - d).days <= 30),
        "last_90d": sum(1 for d in dates if (now - d).days <= 90),
        "last_180d": sum(1 for d in dates if (now - d).days <= 180),
        "last_365d": sum(1 for d in dates if (now - d).days <= 365),
    }


async def _extract_page_meta_single(url: str) -> dict:
    from bs4 import BeautifulSoup
    meta: dict[str, Any] = {
        "url": url, "title": "", "h1": "",
        "meta_description": "", "meta_keywords": "",
        "published_date": "", "canonical": "",
        "word_count_estimate": 0, "status": "ok",
    }
    resp = await _gap_fetch(url, timeout=15, accept="text/html")
    if not resp:
        meta["status"] = "fetch_failed"
        return meta
    try:
        soup = BeautifulSoup(resp.text, "html.parser")
    except Exception:
        meta["status"] = "parse_failed"
        return meta

    t = soup.find("title")
    if t:
        meta["title"] = t.get_text(strip=True)[:200]
    h1 = soup.find("h1")
    if h1:
        meta["h1"] = h1.get_text(strip=True)[:200]
    desc = soup.find("meta", attrs={"name": "description"})
    if desc and desc.get("content"):
        meta["meta_description"] = desc["content"].strip()[:300]
    kw = soup.find("meta", attrs={"name": "keywords"})
    if kw and kw.get("content"):
        meta["meta_keywords"] = kw["content"].strip()[:300]
    can = soup.find("link", attrs={"rel": "canonical"})
    if can and can.get("href"):
        meta["canonical"] = can["href"].strip()

    pub_date = ""
    og_date = soup.find("meta", attrs={"property": "article:published_time"})
    if og_date and og_date.get("content"):
        pub_date = og_date["content"].strip()[:10]
    if not pub_date:
        time_el = soup.find("time")
        if time_el:
            pub_date = (time_el.get("datetime") or time_el.get_text(strip=True))[:10]
    if not pub_date:
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                sd = json.loads(script.string or "")
                if isinstance(sd, dict) and "datePublished" in sd:
                    pub_date = str(sd["datePublished"])[:10]
                    break
                if isinstance(sd, list):
                    for item in sd:
                        if isinstance(item, dict) and "datePublished" in item:
                            pub_date = str(item["datePublished"])[:10]
                            break
            except Exception:
                pass
    meta["published_date"] = pub_date

    body = soup.find("body")
    if body:
        for tag in body(["script", "style", "noscript"]):
            tag.decompose()
        meta["word_count_estimate"] = len(body.get_text(separator=" ", strip=True).split())
    return meta


async def _batch_extract_metadata(urls: list[str], concurrency: int = 10) -> list[dict]:
    sem = asyncio.Semaphore(concurrency)

    async def _limited(u: str):
        async with sem:
            return await _extract_page_meta_single(u)

    return list(await asyncio.gather(*[_limited(u) for u in urls]))


def _build_arch_profile(domain: str, classified: dict,
                        metadata_list: list[dict],
                        all_urls: list[dict]) -> dict:
    now = datetime.now()
    monthly_counts: Counter[str] = Counter()
    recent_content: list[dict] = []

    blog_url_set = {e.get("url") for e in classified.get("blog_posts", [])}

    for m in metadata_list:
        ds = m.get("published_date", "")
        if not ds or len(ds) < 7:
            continue
        ym = ds[:7]
        monthly_counts[ym] += 1
        try:
            d = datetime.strptime(ds[:10], "%Y-%m-%d")
            if (now - d).days <= 90:
                recent_content.append({
                    "url": m["url"], "title": m.get("title", ""),
                    "published_date": ds[:10],
                    "type": "blog_post" if m["url"] in blog_url_set else "landing_page",
                })
        except ValueError:
            pass

    months_12 = [(now - timedelta(days=i * 30)).strftime("%Y-%m") for i in range(11, -1, -1)]
    posts_12 = [monthly_counts.get(ym, 0) for ym in months_12]
    total_12 = sum(posts_12)
    avg = round(total_12 / 12, 1) if total_12 else 0.0

    freq = ("daily" if avg >= 20 else "3x_week" if avg >= 12
            else "2x_week" if avg >= 8 else "weekly" if avg >= 4
            else "bi-weekly" if avg >= 2 else "monthly")

    titles = [m["title"] for m in metadata_list if m.get("title")]
    descs = [m["meta_description"] for m in metadata_list if m.get("meta_description")]
    ml = max(len(metadata_list), 1)

    depths = []
    for m in metadata_list:
        p = urlparse(m.get("url", "")).path
        depths.append(len([s for s in p.strip("/").split("/") if s]))

    recent_content.sort(key=lambda x: x.get("published_date", ""), reverse=True)

    return {
        "domain": domain,
        "publishing_cadence": {
            "posts_last_12_months": posts_12, "months_labels": months_12,
            "avg_per_month": avg, "frequency": freq, "total_last_12_months": total_12,
        },
        "freshness_distribution": _compute_freshness(all_urls),
        "meta_quality": {
            "avg_title_len": round(sum(len(t) for t in titles) / max(len(titles), 1)),
            "avg_desc_len": round(sum(len(d) for d in descs) / max(len(descs), 1)),
            "pct_missing_desc": round((1 - len(descs) / ml) * 100),
            "total_pages_analyzed": len(metadata_list),
        },
        "structural": {"avg_url_depth": round(sum(depths) / max(len(depths), 1), 1)},
        "recent_content": recent_content[:20],
    }


async def _llm_infer_topics(domain: str, metadata_batch: list[dict],
                            model: str) -> dict:
    page_lines = []
    for i, m in enumerate(metadata_batch):
        if m.get("status") != "ok":
            continue
        slug = urlparse(m["url"]).path.rstrip("/").split("/")[-1] if m["url"] else ""
        line = f"{i+1}. URL: {m['url']}"
        if m.get("title"):
            line += f" | Title: {m['title']}"
        if m.get("h1") and m["h1"] != m.get("title"):
            line += f" | H1: {m['h1']}"
        if m.get("meta_description"):
            line += f" | Desc: {m['meta_description'][:100]}"
        if slug:
            line += f" | Slug: {slug}"
        if m.get("published_date"):
            line += f" | Date: {m['published_date']}"
        page_lines.append(line)

    if not page_lines:
        return {"topic_clusters": [], "pages": []}

    batch_size = 50
    all_results: dict[str, list] = {"topic_clusters": [], "pages": []}

    for batch_start in range(0, len(page_lines), batch_size):
        batch = page_lines[batch_start:batch_start + batch_size]

        system_prompt = (
            "You are an expert SEO content strategist. Analyze page metadata and infer:\n"
            "1. Topic clusters (groups of related pages)\n"
            "2. Primary keyword per page\n"
            "3. Search intent per page\n"
            "4. Estimated search volume tier\n\n"
            "Return JSON:\n"
            "{\n"
            '  "topic_clusters": [\n'
            '    {"cluster_name": "...", "pages": [page_numbers], '
            '"primary_theme": "...", "estimated_volume": "high|medium|low"}\n'
            "  ],\n"
            '  "pages": [\n'
            '    {"page_num": 1, "inferred_keyword": "...", '
            '"search_intent": "informational|transactional|commercial|navigational", '
            '"volume_tier": "high|medium|low", '
            '"content_type": "blog_post|landing_page|guide|comparison|listicle|how_to"}\n'
            "  ]\n"
            "}\n\n"
            "RULES:\n"
            "- Infer the most likely target keyword from title, H1, URL slug, description\n"
            "- Group pages into 5-15 logical topic clusters\n"
            "- Volume: high=broad demand, medium=niche, low=very specific/long-tail\n"
            "- Be specific with keywords, not generic descriptions\n"
            "- Every page must appear in exactly one cluster"
        )
        user_prompt = (
            f"Domain: {domain}\n"
            f"Page metadata ({len(batch)} pages):\n"
            + "\n".join(batch)
            + "\n\nAnalyze and return the JSON."
        )

        try:
            r = await llm_review(system_prompt=system_prompt,
                                 user_prompt=user_prompt, model=model)
            if not r.get("parse_error"):
                if batch_start > 0:
                    for c in r.get("topic_clusters", []):
                        c["pages"] = [p + batch_start for p in c.get("pages", [])]
                    for p in r.get("pages", []):
                        p["page_num"] = p.get("page_num", 0) + batch_start
                all_results["topic_clusters"].extend(r.get("topic_clusters", []))
                all_results["pages"].extend(r.get("pages", []))
        except Exception:
            pass

    return all_results


async def _llm_gap_analysis(your_domain: str, your_topics: dict,
                            competitor_data: list[dict], model: str) -> dict:
    your_text = ""
    if your_topics and your_topics.get("topic_clusters"):
        your_text = f"\n--- YOUR SITE ({your_domain}) ---\nTopic Clusters:\n"
        for c in your_topics["topic_clusters"]:
            your_text += (f"- {c['cluster_name']} ({c.get('estimated_volume','?')} vol, "
                          f"{len(c.get('pages',[]))} pages): {c.get('primary_theme','')}\n")
        your_text += "\nPage Keywords:\n"
        for p in your_topics.get("pages", [])[:100]:
            your_text += (f"  - [{p.get('search_intent','?')}] "
                          f"{p.get('inferred_keyword','?')} ({p.get('volume_tier','?')})\n")

    comp_parts = []
    for cd in competitor_data:
        ct = f"\n--- COMPETITOR: {cd['domain']} ---\nTopic Clusters:\n"
        for c in cd.get("topics", {}).get("topic_clusters", []):
            ct += (f"- {c['cluster_name']} ({c.get('estimated_volume','?')} vol, "
                   f"{len(c.get('pages',[]))} pages): {c.get('primary_theme','')}\n")
        ct += "\nPage Keywords:\n"
        for p in cd.get("topics", {}).get("pages", [])[:100]:
            ct += (f"  - [{p.get('search_intent','?')}] "
                   f"{p.get('inferred_keyword','?')} ({p.get('volume_tier','?')})\n")
        comp_parts.append(ct)

    system_prompt = (
        "You are an elite SEO strategist performing content gap analysis.\n\n"
        "Return JSON:\n{\n"
        '  "gap_matrix": [\n'
        '    {"topic":"...","inferred_keyword":"...","search_intent":"informational|transactional|commercial|navigational",'
        '"volume_tier":"high|medium|low","covered_by":["domains"],"your_coverage":"missing|thin|outdated|covered",'
        '"gap_type":"missing|thin|outdated","opportunity_score":1-10,"priority":"high|medium|low"}\n  ],\n'
        '  "recommendations": [\n'
        '    {"suggested_title":"...","content_type":"blog_post|landing_page|guide|comparison|listicle|how_to",'
        '"target_keyword":"...","secondary_keywords":["..."],"search_intent":"...",'
        '"priority_score":1-10,"estimated_effort":"low|medium|high",'
        '"brief_outline":["5-7 sections"],"competing_urls":["..."],"rationale":"..."}\n  ],\n'
        '  "competitor_overlap": {"most_similar_pair":["d1","d2"],"most_differentiated":"d","insights":"..."},\n'
        '  "quick_wins": [{"action":"...","target":"...","expected_impact":"high|medium|low","effort":"low|medium|high"}],\n'
        '  "strategic_moves": [{"action":"...","rationale":"...","expected_impact":"..."}]\n}\n\n'
        "RULES:\n"
        "- Focus on HIGH and MEDIUM volume gaps first\n"
        "- Identify 10-20 content gaps\n"
        "- Generate 10-15 specific recommendations ranked by priority\n"
        "- Be SPECIFIC with keywords\n"
        "- opportunity_score: 10=must-have, 1=nice-to-have\n"
        "- If no user site data, all gaps are 'missing' type"
    )
    user_prompt = (
        "Perform content gap analysis.\n"
        + (your_text or "(No user site — identify all competitor topics as gaps)\n")
        + "\n".join(comp_parts)
        + "\n\nReturn the complete JSON."
    )

    try:
        r = await llm_review(system_prompt=system_prompt,
                             user_prompt=user_prompt, model=model)
        if r.get("parse_error"):
            return {"gap_matrix": [], "recommendations": [],
                    "error": "Failed to parse LLM response"}
        return r
    except Exception as e:
        return {"gap_matrix": [], "recommendations": [], "error": str(e)}


async def _llm_content_calendar(recommendations: list[dict],
                                config: dict, model: str) -> dict:
    freq = config.get("frequency", "weekly")
    horizon = config.get("horizon_months", 3)
    start_date = config.get("start_date", "")

    freq_map = {"daily": 5, "3x_week": 3, "2x_week": 2,
                "weekly": 1, "biweekly": 0.5, "monthly": 0.25}
    ppw = freq_map.get(freq, 1)
    total_weeks = horizon * 4
    total_pieces = int(ppw * total_weeks)

    recs_text = ""
    for i, r in enumerate(recommendations[:30], 1):
        recs_text += (
            f"{i}. Title: {r.get('suggested_title','?')}\n"
            f"   Keyword: {r.get('target_keyword','?')} | Type: {r.get('content_type','?')}\n"
            f"   Intent: {r.get('search_intent','?')} | Priority: {r.get('priority_score','?')}/10"
            f" | Effort: {r.get('estimated_effort','?')}\n"
        )
        if r.get("brief_outline"):
            recs_text += f"   Outline: {' | '.join(r['brief_outline'][:5])}\n"
        recs_text += "\n"

    system_prompt = (
        f"You are a content strategist creating a publishing calendar.\n\n"
        f"Frequency: {freq} ({ppw} pieces/week)\n"
        f"Horizon: {horizon} months (~{total_weeks} weeks)\n"
        f"Start: {start_date or 'next Monday'}\n"
        f"Target: ~{total_pieces} pieces\n\n"
        "Return JSON:\n{\n"
        '  "config": {"frequency":"...","horizon_months":N,"start_date":"YYYY-MM-DD","total_pieces":N},\n'
        '  "monthly_themes": [{"month":"Month Year","theme":"...","pieces":N}],\n'
        '  "calendar": [\n'
        '    {"week":1,"week_start":"YYYY-MM-DD","entries":[\n'
        '      {"publish_date":"YYYY-MM-DD","title":"...","content_type":"blog_post|landing_page",'
        '"target_keyword":"...","search_intent":"...","priority_score":1-10,'
        '"estimated_effort":"low|medium|high","is_pillar":false,'
        '"dependencies":[],"seasonal_note":null,"brief_points":["3-5 points"]}\n'
        "    ]}\n  ],\n"
        '  "milestones": [{"week":N,"label":"..."}],\n'
        '  "summary": {"total_blog_posts":N,"total_landing_pages":N,"high_priority_pieces":N,'
        '"pillar_pages":N,"estimated_total_effort":"..."}\n}\n\n'
        "RULES:\n"
        "- Schedule pillar pages BEFORE cluster posts\n"
        "- Group related topics into 2-3 week sprints\n"
        "- Alternate content types for variety\n"
        "- Mix difficulty levels within each week\n"
        "- Add milestone markers at key points\n"
        "- Each entry needs 3-5 brief_points"
    )
    user_prompt = (
        f"Create a content calendar from these topics:\n\n{recs_text}\n"
        f"Schedule into {horizon} months starting {start_date or 'next Monday'}, "
        f"publishing {freq}. Return JSON."
    )

    try:
        r = await llm_review(system_prompt=system_prompt,
                             user_prompt=user_prompt, model=model)
        if r.get("parse_error"):
            return {"error": "Failed to parse calendar", "calendar": []}
        return r
    except Exception as e:
        return {"error": str(e), "calendar": []}


# ---------------------------------------------------------------------------
# Content Gap API Endpoints
# ---------------------------------------------------------------------------

@app.post("/api/content-gap/discover")
async def content_gap_discover(request: Request):
    form = await request.form()
    domain = form.get("domain", "").strip()
    if not domain:
        raise HTTPException(status_code=400, detail="Domain is required.")

    discovery = await _discover_sitemaps(domain)

    all_urls: list[dict] = []
    sitemap_tree: list[dict] = []
    errors = list(discovery.get("errors", []))

    for sm_url in discovery["sitemap_urls_found"]:
        pr = await _parse_sitemap(sm_url)
        sitemap_tree.extend(pr["sitemap_tree"])
        all_urls.extend(pr["urls"])
        errors.extend(pr["errors"])

    seen: set[str] = set()
    unique: list[dict] = []
    for e in all_urls:
        u = e["url"].rstrip("/")
        if u not in seen:
            seen.add(u)
            unique.append(e)

    if len(unique) > _GAP_MAX_URLS:
        unique.sort(key=lambda x: x.get("lastmod", ""), reverse=True)
        unique = unique[:_GAP_MAX_URLS]

    classified = _classify_gap_urls(unique, domain)
    freshness = _compute_freshness(unique)

    return {
        "stage_1_sitemap": {
            "domain": discovery["domain"],
            "discovery_method": " + ".join(discovery["discovery_method"]) or "none",
            "cms_detected": discovery.get("cms_detected"),
            "sitemap_tree": sitemap_tree,
            "url_summary": {
                "total": classified["total"],
                "blogs": len(classified["blog_posts"]),
                "landing_pages": len(classified["landing_pages"]),
                "other": len(classified["other"]),
            },
            "freshness": freshness,
            "errors": errors,
        }
    }


@app.post("/api/content-gap/analyze")
async def content_gap_analyze(request: Request):
    """Streaming gap analysis — sends SSE events as each stage completes."""
    form = await request.form()
    your_domain = form.get("your_domain", "").strip()
    your_urls_json = form.get("your_urls", "[]")
    comp_domains_json = form.get("competitor_domains", "[]")
    model = form.get("model", "") or DEFAULT_MODEL
    pub_freq = form.get("publishing_frequency", "weekly")
    cal_horizon = form.get("calendar_horizon", "3")
    cal_start = form.get("calendar_start_date", "")

    set_model(model)

    try:
        comp_domains = json.loads(comp_domains_json)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid competitor_domains.")
    try:
        your_urls = json.loads(your_urls_json) if your_urls_json.strip() else []
    except json.JSONDecodeError:
        your_urls = []

    comp_domains = [d.strip() for d in comp_domains if d.strip()]
    if not comp_domains:
        raise HTTPException(status_code=422, detail="Provide at least one competitor domain.")
    if len(comp_domains) > 5:
        raise HTTPException(status_code=422, detail="Maximum 5 competitor domains.")

    def _sse(event: str, data: Any) -> str:
        payload = json.dumps(data, ensure_ascii=False, default=str)
        return f"event: {event}\ndata: {payload}\n\n"

    effective_your_domain = your_domain

    async def _stream():
        nonlocal effective_your_domain
        full_result: dict[str, Any] = {"model_used": model}

        # ── STAGE 1 ──
        log.info("GAP: Stage 1 — Sitemap discovery")
        all_doms: list[tuple[str, str]] = []
        if effective_your_domain:
            all_doms.append(("your", effective_your_domain))
        for cd in comp_domains:
            all_doms.append(("competitor", cd))

        stage_1_data: dict[str, dict] = {}
        for role, dom in all_doms:
            log.info(f"  Discovering {dom}...")
            disc = await _discover_sitemaps(dom)
            raw: list[dict] = []
            stree: list[dict] = []
            errs = list(disc.get("errors", []))
            for sm in disc["sitemap_urls_found"]:
                pr = await _parse_sitemap(sm)
                stree.extend(pr["sitemap_tree"])
                raw.extend(pr["urls"])
                errs.extend(pr["errors"])
            seen: set[str] = set()
            uniq: list[dict] = []
            for e in raw:
                u = e["url"].rstrip("/")
                if u not in seen:
                    seen.add(u)
                    uniq.append(e)
            if len(uniq) > _GAP_MAX_URLS:
                uniq.sort(key=lambda x: x.get("lastmod", ""), reverse=True)
                uniq = uniq[:_GAP_MAX_URLS]
            cl = _classify_gap_urls(uniq, dom)
            log.info(f"  {dom}: {cl['total']} URLs")
            stage_1_data[dom] = {
                "role": role, "domain": disc["domain"],
                "discovery_method": " + ".join(disc["discovery_method"]) or "none",
                "cms_detected": disc.get("cms_detected"), "sitemap_tree": stree,
                "url_summary": {"total": cl["total"], "blogs": len(cl["blog_posts"]),
                                "landing_pages": len(cl["landing_pages"]),
                                "other": len(cl["other"])},
                "freshness": _compute_freshness(uniq),
                "classified": cl, "all_urls": uniq, "errors": errs,
            }

        if your_urls and not effective_your_domain:
            effective_your_domain = "your_site"
            manual = [{"url": u.strip()} for u in your_urls if u.strip()]
            cl = _classify_gap_urls(manual, "your_site")
            stage_1_data["your_site"] = {
                "role": "your", "domain": "your_site",
                "discovery_method": "manual input", "cms_detected": None,
                "sitemap_tree": [], "url_summary": {
                    "total": len(manual), "blogs": len(cl["blog_posts"]),
                    "landing_pages": len(cl["landing_pages"]),
                    "other": len(cl["other"])},
                "freshness": {}, "classified": cl, "all_urls": manual, "errors": [],
            }

        stage_1_out = [{
            "domain": d["domain"], "role": d["role"],
            "discovery_method": d["discovery_method"],
            "cms_detected": d.get("cms_detected"),
            "sitemap_tree": d["sitemap_tree"],
            "url_summary": d["url_summary"],
            "freshness": d.get("freshness", {}),
            "errors": d.get("errors", []),
        } for d in stage_1_data.values()]
        full_result["stage_1_sitemap"] = stage_1_out
        log.info("GAP: Stage 1 complete — sending to client")
        yield _sse("stage_1", {"stage_1_sitemap": stage_1_out})

        # ── STAGE 2 ──
        log.info("GAP: Stage 2 — Metadata extraction")
        stage_2_meta: dict[str, list[dict]] = {}
        stage_2_out = []
        for dom, s1 in stage_1_data.items():
            cl = s1.get("classified", {})
            content_urls = [e["url"] for e in
                            cl.get("blog_posts", []) + cl.get("landing_pages", [])]
            sample = content_urls[:_GAP_MAX_URLS]
            log.info(f"  Extracting metadata for {dom}: {len(sample)} pages...")
            meta = await _batch_extract_metadata(sample, _GAP_METADATA_CONCURRENCY) if sample else []
            ok = sum(1 for m in meta if m.get("status") == "ok")
            log.info(f"  {dom}: {ok}/{len(meta)} extracted OK")
            stage_2_meta[dom] = meta
            stage_2_out.append(_build_arch_profile(s1["domain"], cl, meta, s1.get("all_urls", [])))
        full_result["stage_2_architecture"] = stage_2_out
        log.info("GAP: Stage 2 complete — sending to client")
        yield _sse("stage_2", {"stage_2_architecture": stage_2_out})

        # ── STAGE 3 ──
        log.info("GAP: Stage 3 — LLM topic inference")
        stage_3_data: dict[str, dict] = {}
        stage_3_out = []
        for dom, s1 in stage_1_data.items():
            meta = stage_2_meta.get(dom, [])
            log.info(f"  Topics for {dom} ({len(meta)} pages)...")
            topics = (await _llm_infer_topics(s1["domain"], meta, model)
                      if meta else {"topic_clusters": [], "pages": []})
            log.info(f"  {dom}: {len(topics.get('topic_clusters', []))} clusters")
            stage_3_data[dom] = {"domain": s1["domain"], "role": s1["role"], "topics": topics}
            stage_3_out.append({"domain": s1["domain"], "role": s1["role"],
                                "topic_clusters": topics.get("topic_clusters", []),
                                "pages": topics.get("pages", [])})
        full_result["stage_3_topics"] = stage_3_out
        log.info("GAP: Stage 3 complete — sending to client")
        yield _sse("stage_3", {"stage_3_topics": stage_3_out})

        # ── STAGE 4 & 5 ──
        log.info("GAP: Stage 4+5 — Gap analysis")
        your_topics = None
        comp_list = []
        for dom, s3 in stage_3_data.items():
            if s3["role"] == "your":
                your_topics = s3["topics"]
            else:
                comp_list.append({"domain": s3["domain"], "topics": s3["topics"]})
        gap_r = await _llm_gap_analysis(
            effective_your_domain or "your_site",
            your_topics or {}, comp_list, model,
        )
        log.info(f"GAP: Stage 4+5 complete — {len(gap_r.get('gap_matrix', []))} gaps")
        stage_4_out = {"gap_matrix": gap_r.get("gap_matrix", []),
                       "competitor_overlap": gap_r.get("competitor_overlap", {})}
        stage_5_out = {"recommendations": gap_r.get("recommendations", []),
                       "quick_wins": gap_r.get("quick_wins", []),
                       "strategic_moves": gap_r.get("strategic_moves", [])}
        full_result["stage_4_gaps"] = stage_4_out
        full_result["stage_5_recommendations"] = stage_5_out
        yield _sse("stage_4_5", {"stage_4_gaps": stage_4_out,
                                  "stage_5_recommendations": stage_5_out})

        # ── STAGE 6 ──
        log.info("GAP: Stage 6 — Content calendar")
        cal_cfg = {"frequency": pub_freq,
                   "horizon_months": int(cal_horizon), "start_date": cal_start}
        stage_6_out = await _llm_content_calendar(
            gap_r.get("recommendations", []), cal_cfg, model,
        )
        full_result["stage_6_calendar"] = stage_6_out
        log.info("GAP: Stage 6 complete")
        yield _sse("stage_6", {"stage_6_calendar": stage_6_out})

        # ── COMPLETE ──
        full_result["generated_at"] = datetime.now().isoformat()
        report_id = _save_gap_report(full_result)
        log.info(f"GAP: All stages complete — report {report_id}")
        yield _sse("complete", {"report_id": report_id,
                                 "model_used": model,
                                 "generated_at": full_result["generated_at"]})

    return StreamingResponse(_stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache",
                                      "X-Accel-Buffering": "no"})


def _save_gap_report(data: dict) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    rid = f"gap_{ts}"
    (REPORTS_DIR / f"{rid}.json").write_text(
        json.dumps(data, indent=2, ensure_ascii=False, default=str))
    return rid


def _gap_report_to_text(data: dict) -> str:
    lines = ["=" * 70, "CONTENT GAP ANALYSIS REPORT", "=" * 70, ""]

    lines.append("STAGE 1: SITEMAP INTELLIGENCE")
    lines.append("-" * 50)
    for site in data.get("stage_1_sitemap", []):
        s = site.get("url_summary", {})
        lines.append(f"\n  {site.get('domain','?')} ({site.get('role','?')})")
        lines.append(f"  Discovery: {site.get('discovery_method','?')} | CMS: {site.get('cms_detected','-')}")
        lines.append(f"  Pages: {s.get('total',0)} total | {s.get('blogs',0)} blogs | {s.get('landing_pages',0)} landing")
        f = site.get("freshness", {})
        if f.get("newest_lastmod"):
            lines.append(f"  Latest update: {f['newest_lastmod']} | {f.get('pct_with_lastmod',0)}% dated")

    lines.append("\n\nSTAGE 2: SITE ARCHITECTURE")
    lines.append("-" * 50)
    for p in data.get("stage_2_architecture", []):
        c = p.get("publishing_cadence", {})
        mq = p.get("meta_quality", {})
        lines.append(f"\n  {p.get('domain','?')}: {c.get('avg_per_month',0)} posts/mo ({c.get('frequency','?')})")
        lines.append(f"  Titles avg {mq.get('avg_title_len',0)} chars | Desc avg {mq.get('avg_desc_len',0)} | {mq.get('pct_missing_desc',0)}% missing desc")
        for rc in p.get("recent_content", [])[:5]:
            lines.append(f"    [{rc.get('published_date','?')}] {rc.get('title','?')}")

    lines.append("\n\nSTAGE 3: TOPIC LANDSCAPE")
    lines.append("-" * 50)
    for td in data.get("stage_3_topics", []):
        lines.append(f"\n  {td.get('domain','?')} ({td.get('role','?')})")
        for c in td.get("topic_clusters", []):
            lines.append(f"    - {c.get('cluster_name','?')} ({c.get('estimated_volume','?')} vol, {len(c.get('pages',[]))} pages)")

    lines.append("\n\nSTAGE 4: CONTENT GAPS")
    lines.append("-" * 50)
    for g in data.get("stage_4_gaps", {}).get("gap_matrix", []):
        lines.append(f"  [{g.get('priority','?').upper()}] {g.get('topic','?')} — {g.get('inferred_keyword','?')}")
        lines.append(f"    {g.get('search_intent','?')} | {g.get('volume_tier','?')} vol | {g.get('gap_type','?')} | Score: {g.get('opportunity_score','?')}/10")

    lines.append("\n\nSTAGE 5: RECOMMENDATIONS")
    lines.append("-" * 50)
    for i, r in enumerate(data.get("stage_5_recommendations", {}).get("recommendations", []), 1):
        lines.append(f"\n  {i}. {r.get('suggested_title','?')}")
        lines.append(f"     {r.get('content_type','?')} | {r.get('target_keyword','?')} | Priority: {r.get('priority_score','?')}/10")

    lines.append("\n\nSTAGE 6: CONTENT CALENDAR")
    lines.append("-" * 50)
    cal = data.get("stage_6_calendar", {})
    cfg = cal.get("config", {})
    lines.append(f"  {cfg.get('frequency','?')} | {cfg.get('horizon_months','?')} months | {cfg.get('total_pieces','?')} pieces")
    for w in cal.get("calendar", []):
        lines.append(f"\n  Week {w.get('week','?')} ({w.get('week_start','?')}):")
        for e in w.get("entries", []):
            lines.append(f"    [{e.get('publish_date','?')}] {e.get('title','?')} ({e.get('content_type','?')})")

    lines += ["", f"Model: {data.get('model_used','?')}", f"Generated: {data.get('generated_at','')}", "=" * 70]
    return "\n".join(lines)


@app.post("/api/content-gap/download-text")
async def download_gap_text(request: Request):
    body = await request.json()
    return Response(
        content=_gap_report_to_text(body), media_type="text/plain",
        headers={"Content-Disposition": 'attachment; filename="content_gap_report.txt"'},
    )


@app.post("/api/content-gap/download-excel")
async def download_gap_excel(request: Request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    data = await request.json()
    wb = Workbook()

    PUR = "6D28D9"; WHT = "FFFFFF"; DRK = "1E1E1E"; BDY = "374151"; MUT = "6B7280"
    GRN_BG = "F0FDF4"; ORG_BG = "FFFBEB"; RED_BG = "FEF2F2"; STRIPE = "F3F4F6"
    PUR_LT = "EDE9FE"
    hfont = Font(bold=True, size=10, color=WHT)
    hfill = PatternFill(start_color=PUR, end_color=PUR, fill_type="solid")
    tfont = Font(bold=True, size=13, color=PUR)
    bfont = Font(size=9, color=BDY)
    mfont = Font(size=8, color=MUT)
    sfont = Font(bold=True, size=10, color=DRK)
    ts = Side(style="thin", color="D1D5DB")
    tb = Border(left=ts, right=ts, top=ts, bottom=ts)
    wt = Alignment(wrap_text=True, vertical="top")
    wc = Alignment(wrap_text=True, vertical="center", horizontal="center")

    def mk_hdr(ws, row, hdrs, widths=None):
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hfont; c.fill = hfill; c.border = tb; c.alignment = wc
        if widths:
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

    def pfill(p):
        if isinstance(p, str):
            p = p.lower()
            if p == "high": return PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
            if p == "medium": return PatternFill(start_color=ORG_BG, end_color=ORG_BG, fill_type="solid")
            return PatternFill(start_color=GRN_BG, end_color=GRN_BG, fill_type="solid")
        if isinstance(p, (int, float)):
            if p >= 7: return PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
            if p >= 4: return PatternFill(start_color=ORG_BG, end_color=ORG_BG, fill_type="solid")
            return PatternFill(start_color=GRN_BG, end_color=GRN_BG, fill_type="solid")
        return None

    # Sheet 1: Summary
    ws = wb.active; ws.title = "Executive Summary"; ws.sheet_properties.tabColor = PUR
    ws.merge_cells("A1:E1"); ws["A1"] = "CONTENT GAP ANALYSIS"; ws["A1"].font = tfont
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generated: {data.get('generated_at','')} | Model: {data.get('model_used','')}"
    ws["A2"].font = mfont
    gaps = data.get("stage_4_gaps", {}).get("gap_matrix", [])
    recs = data.get("stage_5_recommendations", {}).get("recommendations", [])
    mk_hdr(ws, 4, ["Metric", "Value"], [35, 20])
    for si, (lbl, val) in enumerate([
        ("Domains Analyzed", len(data.get("stage_1_sitemap", []))),
        ("Content Gaps Found", len(gaps)),
        ("High Priority Gaps", sum(1 for g in gaps if g.get("priority") == "high")),
        ("Recommendations", len(recs)),
    ]):
        r = si + 5
        ws.cell(row=r, column=1, value=lbl).font = sfont
        ws.cell(row=r, column=1).border = tb
        ws.cell(row=r, column=2, value=val).font = bfont
        ws.cell(row=r, column=2).border = tb; ws.cell(row=r, column=2).alignment = wc

    # Sheet 2: Sitemap Intelligence
    ws2 = wb.create_sheet("Sitemap Intelligence"); ws2.sheet_properties.tabColor = "3B82F6"
    ws2.merge_cells("A1:G1"); ws2["A1"] = "SITEMAP INTELLIGENCE"; ws2["A1"].font = tfont
    mk_hdr(ws2, 3, ["Domain", "Role", "CMS", "Discovery", "Total", "Blogs", "Landing"], [25, 12, 18, 22, 10, 10, 12])
    for si, site in enumerate(data.get("stage_1_sitemap", [])):
        r = si + 4; s = site.get("url_summary", {})
        for ci, v in enumerate([site.get("domain"), site.get("role"), site.get("cms_detected", "-"),
                                site.get("discovery_method"), s.get("total", 0),
                                s.get("blogs", 0), s.get("landing_pages", 0)], 1):
            c = ws2.cell(row=r, column=ci, value=v); c.font = bfont; c.border = tb; c.alignment = wt

    # Sheet 3: Architecture
    ws3 = wb.create_sheet("Architecture"); ws3.sheet_properties.tabColor = "22C55E"
    ws3.merge_cells("A1:F1"); ws3["A1"] = "SITE ARCHITECTURE"; ws3["A1"].font = tfont
    mk_hdr(ws3, 3, ["Domain", "Posts/Mo", "Frequency", "Avg Title", "Avg Desc", "Missing Desc %"], [25, 12, 14, 12, 12, 14])
    for si, prof in enumerate(data.get("stage_2_architecture", [])):
        r = si + 4; cad = prof.get("publishing_cadence", {}); mq = prof.get("meta_quality", {})
        for ci, v in enumerate([prof.get("domain"), cad.get("avg_per_month", 0), cad.get("frequency", "?"),
                                mq.get("avg_title_len", 0), mq.get("avg_desc_len", 0),
                                mq.get("pct_missing_desc", 0)], 1):
            c = ws3.cell(row=r, column=ci, value=v); c.font = bfont; c.border = tb

    # Sheet 4: Topics
    ws4 = wb.create_sheet("Topic Landscape"); ws4.sheet_properties.tabColor = "F59E0B"
    ws4.merge_cells("A1:E1"); ws4["A1"] = "TOPIC LANDSCAPE"; ws4["A1"].font = tfont
    mk_hdr(ws4, 3, ["Domain", "Cluster", "Theme", "Volume", "Pages"], [20, 25, 35, 12, 8])
    r4 = 3
    for td in data.get("stage_3_topics", []):
        for cl in td.get("topic_clusters", []):
            r4 += 1
            for ci, v in enumerate([td.get("domain"), cl.get("cluster_name"), cl.get("primary_theme"),
                                    cl.get("estimated_volume"), len(cl.get("pages", []))], 1):
                c = ws4.cell(row=r4, column=ci, value=v); c.font = bfont; c.border = tb; c.alignment = wt

    # Sheet 5: Gaps
    ws5 = wb.create_sheet("Content Gaps"); ws5.sheet_properties.tabColor = "EF4444"
    ws5.merge_cells("A1:H1"); ws5["A1"] = "CONTENT GAP MATRIX"; ws5["A1"].font = tfont
    mk_hdr(ws5, 3, ["Topic", "Keyword", "Intent", "Volume", "Gap", "Covered By", "Score", "Priority"],
           [25, 22, 14, 10, 10, 25, 8, 10])
    for gi, g in enumerate(gaps):
        r = gi + 4
        vals = [g.get("topic"), g.get("inferred_keyword"), g.get("search_intent"),
                g.get("volume_tier"), g.get("gap_type"), ", ".join(g.get("covered_by", [])),
                g.get("opportunity_score"), g.get("priority")]
        for ci, v in enumerate(vals, 1):
            c = ws5.cell(row=r, column=ci, value=v); c.font = bfont; c.border = tb; c.alignment = wt
            if ci in (7, 8):
                pf = pfill(v)
                if pf: c.fill = pf

    # Sheet 6: Recommendations
    ws6 = wb.create_sheet("Recommendations"); ws6.sheet_properties.tabColor = PUR
    ws6.merge_cells("A1:H1"); ws6["A1"] = "RECOMMENDATIONS"; ws6["A1"].font = tfont
    mk_hdr(ws6, 3, ["#", "Title", "Type", "Keyword", "Intent", "Priority", "Effort", "Outline"],
           [4, 30, 12, 20, 12, 8, 10, 40])
    for ri, rec in enumerate(recs):
        r = ri + 4
        ol = " > ".join(rec.get("brief_outline", []))
        for ci, v in enumerate([ri + 1, rec.get("suggested_title"), rec.get("content_type"),
                                rec.get("target_keyword"), rec.get("search_intent"),
                                rec.get("priority_score"), rec.get("estimated_effort"), ol], 1):
            c = ws6.cell(row=r, column=ci, value=v); c.font = bfont; c.border = tb; c.alignment = wt

    # Sheet 7: Calendar
    ws7 = wb.create_sheet("Content Calendar"); ws7.sheet_properties.tabColor = "22C55E"
    ws7.merge_cells("A1:H1"); ws7["A1"] = "CONTENT CALENDAR"; ws7["A1"].font = tfont
    cal = data.get("stage_6_calendar", {}); cfg = cal.get("config", {})
    ws7.merge_cells("A2:H2")
    ws7["A2"] = f"Freq: {cfg.get('frequency','?')} | {cfg.get('horizon_months','?')} months | Start: {cfg.get('start_date','?')}"
    ws7["A2"].font = mfont
    mk_hdr(ws7, 4, ["Week", "Date", "Title", "Type", "Keyword", "Intent", "Priority", "Effort"],
           [6, 12, 30, 12, 20, 12, 8, 10])
    r7 = 4
    for w in cal.get("calendar", []):
        for e in w.get("entries", []):
            r7 += 1
            for ci, v in enumerate([w.get("week"), e.get("publish_date"), e.get("title"),
                                    e.get("content_type"), e.get("target_keyword"),
                                    e.get("search_intent"), e.get("priority_score"),
                                    e.get("estimated_effort")], 1):
                c = ws7.cell(row=r7, column=ci, value=v); c.font = bfont; c.border = tb; c.alignment = wt

    for sheet in [ws, ws2, ws3, ws4, ws5, ws6, ws7]:
        sheet.freeze_panes = "A5" if sheet in (ws, ws7) else "A4"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="content_gap_report.xlsx"'},
    )


@app.post("/api/content-gap/download-pdf")
async def download_gap_pdf(request: Request):
    from fpdf import FPDF
    import io

    data = await request.json()
    S = _sanitize_for_pdf
    PUR = (109, 40, 217); PUR_L = (237, 233, 254)
    DRK = (30, 30, 30); BDY = (55, 55, 55); MUT = (120, 120, 120)
    WHT = (255, 255, 255); GRN = (34, 197, 94); ORG = (245, 158, 11)
    RED = (239, 68, 68); BRD = (209, 213, 219); STR = (243, 244, 246)

    class PDF(FPDF):
        def header(self):
            if self.page_no() == 1: return
            self.set_font("Helvetica", "", 7); self.set_text_color(*MUT)
            self.set_y(8); self.cell(0, 5, "Vizup Soul | Content Gap Analysis", align="R")
            self.set_draw_color(*PUR); self.line(10, 14, 200, 14); self.set_y(17)
        def footer(self):
            self.set_y(-12); self.set_font("Helvetica", "", 7); self.set_text_color(*MUT)
            self.cell(0, 5, f"Page {self.page_no()}/{{nb}}", align="C")

    pdf = PDF("P", "mm", "A4"); pdf.alias_nb_pages()
    pdf.set_auto_page_break(True, 16)
    LM = pdf.l_margin; PW = pdf.w - pdf.l_margin - pdf.r_margin

    def sec(title):
        if pdf.get_y() > pdf.h - 35: pdf.add_page()
        pdf.ln(3); pdf.set_fill_color(*PUR)
        pdf.rect(LM, pdf.get_y(), PW, 9, "F")
        pdf.set_font("Helvetica", "B", 11); pdf.set_text_color(*WHT)
        pdf.set_x(LM + 4); pdf.cell(PW - 8, 9, S(title), new_x="LMARGIN", new_y="NEXT"); pdf.ln(3)

    # Cover
    pdf.add_page(); pdf.set_fill_color(*PUR); pdf.rect(0, 0, 210, 48, "F")
    pdf.set_y(10); pdf.set_font("Helvetica", "B", 22); pdf.set_text_color(*WHT)
    pdf.cell(0, 12, "Content Gap Analysis", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10); pdf.set_text_color(220, 210, 255)
    pdf.cell(0, 6, "Vizup Soul", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 5, S(f"Generated: {data.get('generated_at','')} | Model: {data.get('model_used','')}"),
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_y(55)

    # Stage 1
    sec("Stage 1: Sitemap Intelligence")
    for site in data.get("stage_1_sitemap", []):
        pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(*PUR)
        pdf.cell(0, 5, S(f"{site.get('domain','?')} ({site.get('role','?')})"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*BDY)
        s = site.get("url_summary", {})
        pdf.cell(0, 4, S(f"Discovery: {site.get('discovery_method','?')} | CMS: {site.get('cms_detected','-')}"), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 4, S(f"{s.get('total',0)} pages | {s.get('blogs',0)} blogs | {s.get('landing_pages',0)} landing"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    # Stage 2
    sec("Stage 2: Architecture")
    for prof in data.get("stage_2_architecture", []):
        cad = prof.get("publishing_cadence", {}); mq = prof.get("meta_quality", {})
        pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(*PUR)
        pdf.cell(0, 5, S(prof.get("domain", "?")), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*BDY)
        pdf.cell(0, 4, S(f"{cad.get('avg_per_month',0)} posts/mo ({cad.get('frequency','?')}) | Titles avg {mq.get('avg_title_len',0)} chars"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    # Stage 4
    gaps = data.get("stage_4_gaps", {}).get("gap_matrix", [])
    if gaps:
        sec("Stage 4: Content Gaps")
        for i, g in enumerate(gaps):
            if pdf.get_y() > pdf.h - 16: pdf.add_page()
            pdf.set_font("Helvetica", "B", 8); pdf.set_text_color(*DRK)
            pdf.cell(0, 5, S(f"{i+1}. {g.get('topic','?')} -- {g.get('inferred_keyword','?')}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 7); pdf.set_text_color(*BDY)
            pdf.cell(0, 4, S(f"{g.get('search_intent','?')} | {g.get('volume_tier','?')} | {g.get('gap_type','?')} | Score: {g.get('opportunity_score','?')}/10"), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

    # Stage 5
    recs = data.get("stage_5_recommendations", {}).get("recommendations", [])
    if recs:
        sec("Stage 5: Recommendations")
        for i, r in enumerate(recs):
            if pdf.get_y() > pdf.h - 20: pdf.add_page()
            pdf.set_font("Helvetica", "B", 8); pdf.set_text_color(*PUR)
            pdf.cell(0, 5, S(f"{i+1}. {r.get('suggested_title','?')}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 7); pdf.set_text_color(*BDY)
            pdf.cell(0, 4, S(f"{r.get('content_type','?')} | {r.get('target_keyword','?')} | P{r.get('priority_score','?')}/10"), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

    # Stage 6
    cal = data.get("stage_6_calendar", {})
    if cal.get("calendar"):
        sec("Stage 6: Content Calendar")
        cfg = cal.get("config", {})
        pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*BDY)
        pdf.cell(0, 5, S(f"{cfg.get('frequency','?')} | {cfg.get('horizon_months','?')} months | {cfg.get('total_pieces','?')} pieces"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        for w in cal.get("calendar", []):
            if pdf.get_y() > pdf.h - 18: pdf.add_page()
            pdf.set_font("Helvetica", "B", 8); pdf.set_text_color(*DRK)
            pdf.cell(0, 5, S(f"Week {w.get('week','?')} -- {w.get('week_start','?')}"), new_x="LMARGIN", new_y="NEXT")
            for e in w.get("entries", []):
                pdf.set_font("Helvetica", "", 7); pdf.set_text_color(*BDY); pdf.set_x(LM + 4)
                pdf.cell(0, 4, S(f"[{e.get('publish_date','?')}] {e.get('title','?')} ({e.get('content_type','?')})"), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

    buf = io.BytesIO(); pdf.output(buf); buf.seek(0)
    return Response(
        content=buf.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="content_gap_report.pdf"'},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8500)
